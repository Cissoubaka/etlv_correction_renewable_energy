#!/usr/bin/env python3
"""Script de test pour analyser la structure du fichier Calc/Excel"""

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from odf import opendocument, table, text
    ODFPY_AVAILABLE = True
except ImportError:
    ODFPY_AVAILABLE = False

def analyze_excel(file_path):
    """Analyse la structure Excel"""
    print(f"\n📊 Analyse du fichier Excel: {file_path}\n")
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    print(f"Feuille: {ws.title}")
    print(f"Dimensions: {ws.dimensions}\n")
    
    # Afficher les 10 premières lignes et 10 premières colonnes
    print("Structure du tableau:\n")
    for row in range(1, min(11, ws.max_row + 1)):
        print(f"Ligne {row}: ", end="")
        for col in range(1, min(11, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            value = cell.value
            if value:
                print(f"[{value}] ", end="")
            else:
                print("[ ] ", end="")
        print()

def analyze_ods(file_path):
    """Analyse la structure ODS"""
    print(f"\n📊 Analyse du fichier ODS: {file_path}\n")
    
    doc = opendocument.load(file_path)
    sheets = doc.spreadsheet.getElementsByType(table.Table)
    
    if not sheets:
        print("Aucune feuille trouvée!")
        return
    
    sheet = sheets[0]
    sheet_name = sheet.getAttribute('name')
    print(f"Feuille: {sheet_name}\n")
    
    rows = sheet.getElementsByType(table.TableRow)
    
    # Afficher les 10 premières lignes
    print("Structure du tableau:\n")
    for row_idx in range(min(10, len(rows))):
        row = rows[row_idx]
        cells = row.getElementsByType(table.TableCell)
        
        print(f"Ligne {row_idx + 1}: ", end="")
        for col_idx, cell in enumerate(cells[:10]):  # Afficher les 10 premières colonnes
            text_elements = cell.getElementsByType(text.P)
            cell_text = "".join([str(e) for e in text_elements]) if text_elements else ""
            
            if cell_text:
                # Truncate long text
                display_text = cell_text[:15] if len(cell_text) > 15 else cell_text
                print(f"[{display_text}] ", end="")
            else:
                print("[ ] ", end="")
        print()

def main():
    if len(sys.argv) < 2:
        print("Usage: python analyze_spreadsheet.py <file.xlsx|file.ods>")
        sys.exit(1)
    
    file_path = Path(sys.argv[1])
    
    if not file_path.exists():
        print(f"❌ Fichier non trouvé: {file_path}")
        sys.exit(1)
    
    try:
        if file_path.suffix.lower() == '.xlsx':
            if not OPENPYXL_AVAILABLE:
                print("❌ openpyxl n'est pas installé")
                sys.exit(1)
            analyze_excel(file_path)
        elif file_path.suffix.lower() in ['.ods', '.calc']:
            if not ODFPY_AVAILABLE:
                print("❌ odfpy n'est pas installé")
                sys.exit(1)
            analyze_ods(file_path)
        else:
            print(f"❌ Format non supporté: {file_path.suffix}")
            sys.exit(1)
    
    except Exception as e:
        print(f"❌ Erreur: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
