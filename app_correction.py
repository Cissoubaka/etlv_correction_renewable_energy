#!/usr/bin/env python3
"""
Application graphique pour corriger rapidement les formulaires PDF
Compare les worksheets des étudiants avec les fichiers de correction
"""

import os
import sys
import json
import re
import hashlib
from pathlib import Path
from datetime import datetime, date
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QFileDialog, QLabel, QListWidget, QListWidgetItem,
    QSplitter, QTableWidget, QTableWidgetItem, QHeaderView, QMessageBox,
    QDialog, QCalendarWidget, QSpinBox, QComboBox, QSpinBox, QGroupBox,
    QTextEdit, QFrame, QCheckBox
)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QColor, QFont
import pdfplumber
from pdf2image import convert_from_path
from PIL import Image
import io
from PyQt5.QtWidgets import QScrollArea, QGridLayout
from PyQt5.QtGui import QPixmap, QImage
from PyQt5.QtCore import QSize

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from odf import opendocument, table
    ODFPY_AVAILABLE = True
except ImportError:
    ODFPY_AVAILABLE = False

try:
    from PyPDF2 import PdfReader
    PYPDF2_AVAILABLE = True
except ImportError:
    PYPDF2_AVAILABLE = False

import subprocess
import platform

class ClickableLabel(QLabel):
    """Label cliquable qui ouvre un fichier PDF"""
    def __init__(self, pdf_path=None, parent=None):
        super().__init__(parent)
        self.pdf_path = pdf_path
        self.setCursor(Qt.PointingHandCursor if pdf_path else Qt.ArrowCursor)
    
    def set_pdf_path(self, pdf_path):
        """Définit le chemin du PDF à ouvrir"""
        self.pdf_path = pdf_path
        self.setCursor(Qt.PointingHandCursor if pdf_path else Qt.ArrowCursor)
    
    def mousePressEvent(self, event):
        """Ouvre le PDF au clic"""
        if self.pdf_path and Path(self.pdf_path).exists():
            try:
                pdf_path = Path(self.pdf_path)
                if platform.system() == 'Darwin':  # macOS
                    subprocess.Popen(['open', str(pdf_path)])
                elif platform.system() == 'Windows':
                    os.startfile(str(pdf_path))
                else:  # Linux
                    subprocess.Popen(['xdg-open', str(pdf_path)])
            except Exception as e:
                QMessageBox.warning(None, "Erreur", f"Impossible d'ouvrir le PDF:\n{str(e)}")
        else:
            super().mousePressEvent(event)

class FieldTextDialog(QDialog):
    """Dialogue pour afficher le texte complet d'un champ"""
    def __init__(self, field_name, student_value, correction_value, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Contenu - {field_name}")
        self.setGeometry(100, 100, 600, 400)
        
        layout = QVBoxLayout()
        
        # Titre
        title = QLabel(f"<b>Champ: {field_name}</b>")
        layout.addWidget(title)
        
        # Réponse étudiant
        layout.addWidget(QLabel("<b>Réponse étudiant:</b>"))
        student_text = QTextEdit()
        student_text.setPlainText(student_value or "[VIDE]")
        student_text.setReadOnly(True)
        if student_value and student_value.strip():
            student_text.setStyleSheet("background-color: #E8F5E9;")  # Vert clair
        else:
            student_text.setStyleSheet("background-color: #FFEBEE;")  # Rouge clair
        layout.addWidget(student_text)
        
        # Réponse correction
        layout.addWidget(QLabel("<b>Réponse correction:</b>"))
        correction_text = QTextEdit()
        correction_text.setPlainText(correction_value or "[VIDE]")
        correction_text.setReadOnly(True)
        correction_text.setStyleSheet("background-color: #E8F5E9;")  # Vert clair
        layout.addWidget(correction_text)
        
        # Bouton fermer
        btn_close = QPushButton("Fermer")
        btn_close.clicked.connect(self.accept)
        layout.addWidget(btn_close)
        
        self.setLayout(layout)

class SpreadsheetParser:
    """Parse les fichiers Calc/Excel pour extraire les dates de deadlines"""
    
    # Durée en heures pour chaque worksheet
    WORKSHEET_DURATION = {
        1: 2,  # 2 heures
        2: 2,  # 2 heures
        3: 1,  # 1 heure
        4: 1,  # 1 heure
    }
    
    @staticmethod
    def parse_spreadsheet(file_path):
        """
        Extrait les dates des worksheets depuis un fichier Calc/Excel.
        Retourne un dictionnaire: {student_name: {1: date, 2: date, ...}, ...}
        ou {1: date, 2: date, ...} si format global
        """
        file_path = Path(file_path)
        
        if file_path.suffix.lower() == '.xlsx':
            return SpreadsheetParser._parse_excel(file_path)
        elif file_path.suffix.lower() in ['.ods', '.calc']:
            return SpreadsheetParser._parse_ods(file_path)
        else:
            raise ValueError(f"Format de fichier non supporté: {file_path.suffix}")
    
    @staticmethod
    def _parse_excel(file_path):
        """Parse un fichier Excel (.xlsx) avec le format de présence"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl n'est pas installé")
        
        try:
            wb = load_workbook(file_path)
            
            # Chercher l'onglet "Presence"
            ws = None
            for sheet in wb.sheetnames:
                if "presence" in sheet.lower():
                    ws = wb[sheet]
                    break
            
            if not ws:
                ws = wb.active
            
            print(f"  📄 Feuille utilisée: {ws.title}")
            
            return SpreadsheetParser._parse_presence_sheet_excel(ws)
        
        except Exception as e:
            print(f"❌ Erreur lors de la lecture du fichier Excel: {e}")
            import traceback
            traceback.print_exc()
            return {}
    
    @staticmethod
    def _parse_ods(file_path):
        """Parse un fichier ODS (.ods) avec le format de présence"""
        if not ODFPY_AVAILABLE:
            raise ImportError("odfpy n'est pas installé")
        
        try:
            doc = opendocument.load(file_path)
            sheets = doc.spreadsheet.getElementsByType(table.Table)
            
            # Chercher l'onglet "Presence"
            ws = None
            for sheet in sheets:
                if "presence" in sheet.getAttribute('name').lower():
                    ws = sheet
                    break
            
            if not ws:
                ws = sheets[0] if sheets else None
            
            if not ws:
                return {}
            
            print(f"  📄 Feuille utilisée: {ws.getAttribute('name')}")
            
            return SpreadsheetParser._parse_presence_sheet_ods(ws)
        
        except Exception as e:
            print(f"❌ Erreur lors de la lecture du fichier ODS: {e}")
            import traceback
            traceback.print_exc()
            return {}
    
    @staticmethod
    def _parse_presence_sheet_excel(worksheet):
        """Parse la feuille de présence Excel pour extraire les dates par élève"""
        student_deadlines = {}
        
        # Lire la ligne 7 pour les dates des séances
        dates_row = 7
        session_dates = {}  # {colonne: date}
        
        print(f"  📅 Lecture des dates (ligne {dates_row})...")
        
        for col in range(2, 20):  # Colonnes B à S (2 à 19)
            cell = worksheet.cell(row=dates_row, column=col)
            if not cell.value:
                continue
            
            date_value = SpreadsheetParser._parse_date_value(cell.value)
            if date_value:
                session_dates[col] = date_value
                print(f"    - Colonne {chr(64+col)}: {date_value}")
        
        if not session_dates:
            print("  ⚠️ Aucune date trouvée en ligne 7")
            return {}
        
        # Lire les élèves et leurs activités à partir de la ligne 8
        print(f"  👥 Lecture des élèves...")
        
        for row in range(8, 60):  # À partir de la ligne 8
            # Colonne A = nom de l'élève
            name_cell = worksheet.cell(row=row, column=1)
            if not name_cell.value:
                continue
            
            student_name = str(name_cell.value).strip()
            if not student_name or student_name.upper() == "ABS":
                continue
            
            print(f"    - Élève: {student_name}")
            student_deadlines[student_name] = {}
            
            # Pour chaque colonne (séance), vérifier l'activité
            for col, session_date in session_dates.items():
                activity_cell = worksheet.cell(row=row, column=col)
                if not activity_cell.value:
                    continue
                
                activity_value = str(activity_cell.value).strip()
                
                # Vérifier si c'est un numéro de worksheet (simple entier)
                try:
                    ws_num = int(activity_value)
                    if 1 <= ws_num <= 4:
                        # Garder la DERNIÈRE date pour ce worksheet (overwrite)
                        student_deadlines[student_name][ws_num] = session_date
                        print(f"      ✓ Worksheet {ws_num}: {session_date}")
                
                except (ValueError, AttributeError):
                    pass
        
        return student_deadlines
    
    @staticmethod
    def _parse_presence_sheet_ods(worksheet):
        """Parse la feuille de présence ODS - ligne 7=dates, lignes 8+=élèves"""
        student_deadlines = {}
        
        # Convertir en liste de lignes
        rows = worksheet.getElementsByType(table.TableRow)
        
        if len(rows) < 8:
            print("  ⚠️ Le fichier n'a pas assez de lignes")
            return {}
        
        # Lire la ligne 7 (index 6) pour les dates des séances
        dates_row = rows[6] if len(rows) > 6 else None
        session_dates = {}  # {col_index: date}
        
        if dates_row:
            print(f"  📅 Lecture des dates (ligne 7)...")
            cells = dates_row.getElementsByType(table.TableCell)
            
            # Parcourir à partir de colonne B (index 1)
            for i, cell in enumerate(cells[1:]):  # i=0 pour cells[1]=colonne B
                col_index = i + 1  # col_index=1 pour cells[1], col_index=2 pour cells[2]
                cell_text = SpreadsheetParser._extract_ods_cell_text(cell)
                if cell_text and cell_text.strip():
                    date_value = SpreadsheetParser._parse_date_string(cell_text)
                    if date_value:
                        session_dates[col_index] = date_value
                        col_letter = chr(65 + col_index)  # A=65, donc B=66, C=67, etc.
                        print(f"    - Colonne {col_letter}: {date_value}")
        
        if not session_dates:
            print("  ⚠️ Aucune date trouvée en ligne 7")
            return {}
        
        # Lire les élèves à partir de la ligne 8 (index 7)
        print(f"  👥 Lecture des élèves...")
        
        for row_idx in range(7, min(len(rows), 60)):  # Lignes 8+ (indices 7+)
            row = rows[row_idx]
            cells = row.getElementsByType(table.TableCell)
            
            if not cells:
                continue
            
            # Première cellule = nom de l'élève
            name_text = SpreadsheetParser._extract_ods_cell_text(cells[0])
            if not name_text or name_text.upper() == "ABS":
                continue
            
            student_name = name_text.strip()
            print(f"    - Élève: {student_name}")
            student_deadlines[student_name] = {}
            
            # Pour chaque colonne (séance)
            for col_index, session_date in session_dates.items():
                if col_index >= len(cells):
                    continue
                
                activity_text = SpreadsheetParser._extract_ods_cell_text(cells[col_index])
                if not activity_text:
                    continue
                
                activity_value = activity_text.strip()
                
                # Vérifier si c'est un numéro de worksheet (simple entier)
                try:
                    ws_num = int(activity_value)
                    if 1 <= ws_num <= 4:
                        # Garder la DERNIÈRE date pour ce worksheet (overwrite)
                        student_deadlines[student_name][ws_num] = session_date
                        print(f"      ✓ Worksheet {ws_num}: {session_date}")
                
                except (ValueError, AttributeError):
                    pass
        
        return student_deadlines
    
    @staticmethod
    def _extract_ods_cell_text(cell):
        """Extrait le texte d'une cellule ODS"""
        try:
            from odf import text
            text_elements = cell.getElementsByType(text.P)
            if text_elements:
                return "".join([str(e) for e in text_elements])
            return ""
        except:
            return ""
    
    @staticmethod
    def _parse_date_value(value):
        """Parse une valeur de date (Excel ou chaîne)"""
        if value is None:
            return None
        
        if isinstance(value, date):
            return value
        
        if isinstance(value, datetime):
            return value.date()
        
        # Essayer de parser comme chaîne
        return SpreadsheetParser._parse_date_string(str(value))
    
    @staticmethod
    def _parse_date_string(date_str):
        """Parse une chaîne de date dans différents formats"""
        if not date_str or not isinstance(date_str, str):
            return None
        
        date_str = date_str.strip()
        
        # Formats courants
        formats = [
            "%d/%m/%y",      # 17/10/25
            "%d/%m/%Y",      # 17/10/2025
            "%Y-%m-%d",      # 2025-10-17
            "%d-%m-%Y",      # 17-10-2025
            "%d.%m.%Y",      # 17.10.2025
            "%B %d, %Y",     # October 17, 2025
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        
        return None

class ConfigManager:
    """Gère la configuration de l'application (chemins, deadlines)"""
    CONFIG_FILE = Path.home() / ".correction_app" / "config.json"
    
    # Patterns pour trouver les worksheets
    WORKSHEET_PATTERNS = {
        1: [
            r"worksheet\s*1(?:\s|\.|\-|_|$)", 
            r"worksheet\s+one(?:\s|\.|\-|_|$)", 
            r"worksheet\s*one\b",
            r"activité\s*1(?:\s|\.|\-|_|$)",
            r"act\s+1(?:\s|\.|\-|_|$)",
            r"^.*worksheet.*1.*\.pdf$",
            r"^.*activité.*1.*\.pdf$",
        ],
        2: [
            r"worksheet\s*2(?:\s|\.|\-|_|$)",
            r"worksheet\s+two(?:\s|\.|\-|_|$)",
            r"worksheet\s*two\b",
            r"activité\s*2(?:\s|\.|\-|_|$)",
            r"act\s+2(?:\s|\.|\-|_|$)",
            r"^.*worksheet.*2.*\.pdf$",
            r"^.*activité.*2.*\.pdf$",
        ],
        3: [
            r"worksheet\s*3(?:\s|\.|\-|_|$)",
            r"worksheet\s+three(?:\s|\.|\-|_|$)",
            r"worksheet\s*three\b",
            r"activité\s*3(?:\s|\.|\-|_|$)",
            r"act\s+3(?:\s|\.|\-|_|$)",
            r"^.*worksheet.*3.*\.pdf$",
            r"^.*activité.*3.*\.pdf$",
        ],
        4: [
            r"worksheet\s*4(?:\s|\.|\-|_|$)",
            r"worksheet\s+four(?:\s|\.|\-|_|$)",
            r"worksheet\s*four\b",
            r"activité\s*4(?:\s|\.|\-|_|$)",
            r"act\s+4(?:\s|\.|\-|_|$)",
            r"^.*worksheet.*4.*\.pdf$",
            r"^.*activité.*4.*\.pdf$",
        ],
    }
    
    def __init__(self):
        self.CONFIG_FILE.parent.mkdir(parents=True, exist_ok=True)
        self.config = self.load_config()
    
    def load_config(self):
        """Charge la configuration depuis le fichier JSON"""
        if self.CONFIG_FILE.exists():
            try:
                with open(self.CONFIG_FILE, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    # Ajouter les clés manquantes si elles n'existent pas
                    if "manual_selections" not in config:
                        config["manual_selections"] = {}
                    return config
            except Exception as e:
                print(f"Erreur lors de la lecture de la config: {e}")
                return self._default_config()
        return self._default_config()
    
    def _default_config(self):
        """Retourne la configuration par défaut"""
        return {
            "work_dir": "",
            "correction_dir": "",
            "deadlines": {
                "worksheet1": None,
                "worksheet2": None,
                "worksheet3": None,
                "worksheet4": None
            },
            "manual_selections": {}
        }
    
    def save_config(self):
        """Sauvegarde la configuration dans le fichier JSON"""
        try:
            with open(self.CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Erreur lors de la sauvegarde de la config: {e}")
    
    def set_work_dir(self, path):
        """Définit le répertoire de travail"""
        self.config["work_dir"] = str(path)
        self.save_config()
    
    def set_correction_dir(self, path):
        """Définit le répertoire des corrections"""
        self.config["correction_dir"] = str(path)
        self.save_config()
    
    def set_deadline(self, worksheet_num, deadline_date):
        """Définit la deadline pour un worksheet"""
        ws_name = f"worksheet{worksheet_num}"
        self.config["deadlines"][ws_name] = str(deadline_date) if deadline_date else None
        self.save_config()
    
    def get_deadline(self, worksheet_num):
        """Récupère la deadline pour un worksheet"""
        ws_name = f"worksheet{worksheet_num}"
        deadline_str = self.config["deadlines"].get(ws_name)
        if deadline_str:
            try:
                return datetime.strptime(deadline_str, "%Y-%m-%d").date()
            except:
                return None
        return None
    
    def save_manual_selection(self, student_name, worksheet_num, file_path):
        """Sauvegarde un fichier sélectionné manuellement"""
        if student_name not in self.config["manual_selections"]:
            self.config["manual_selections"][student_name] = {}
        
        self.config["manual_selections"][student_name][f"worksheet{worksheet_num}"] = str(file_path)
        self.save_config()
    
    def get_manual_selection(self, student_name, worksheet_num):
        """Récupère un fichier sélectionné manuellement"""
        if "manual_selections" not in self.config:
            self.config["manual_selections"] = {}
        
        if student_name not in self.config["manual_selections"]:
            return None
        
        file_path_str = self.config["manual_selections"][student_name].get(f"worksheet{worksheet_num}")
        if file_path_str:
            file_path = Path(file_path_str)
            if file_path.exists():
                return file_path
            else:
                # Supprimer l'entrée si le fichier n'existe plus
                del self.config["manual_selections"][student_name][f"worksheet{worksheet_num}"]
                self.save_config()
                return None
        return None
    
    @staticmethod
    def find_worksheet_file(folder_path, ws_num):
        """
        Cherche un fichier worksheet dans un dossier avec différentes variantes de nom.
        Retourne le chemin du fichier PDF le plus probable, ou None.
        """
        folder_path = Path(folder_path)
        patterns = ConfigManager.WORKSHEET_PATTERNS.get(ws_num, [])
        
        # Tous les fichiers PDF du dossier et ses sous-dossiers
        pdf_files = list(folder_path.rglob("*.pdf"))
        
        # Chercher avec les patterns (dans l'ordre, du plus au moins spécifique)
        for pattern in patterns:
            for pdf_file in pdf_files:
                if re.search(pattern, pdf_file.name, re.IGNORECASE):
                    return pdf_file
        
        return None


class CheatDetectionDialog(QDialog):
    """Dialog pour détecter les fraudes en comparant les hashes des réponses"""
    
    def __init__(self, students_dict, worksheet_num, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Détection de Fraude - Worksheet {worksheet_num}")
        self.setGeometry(100, 100, 1200, 600)
        
        self.students_dict = students_dict
        self.worksheet_num = worksheet_num
        
        self.init_ui()
        self.analyze_cheating()
    
    def _calculate_answers_hash(self, pdf_path):
        """Calcule un hash basé sur toutes les réponses du PDF"""
        try:
            if not PYPDF2_AVAILABLE:
                return None
            
            reader = PdfReader(str(pdf_path))
            if not reader.get_fields():
                return None
            
            # Concaténer toutes les réponses
            all_answers = []
            for field_name, field_obj in reader.get_fields().items():
                value = field_obj.get('/V')
                if value:
                    if isinstance(value, bytes):
                        try:
                            value = value.decode('utf-8')
                        except Exception:
                            value = str(value)
                    else:
                        value = str(value)
                all_answers.append(f"{field_name}:{value}")
            
            # Trier pour assurer une cohérence
            combined = "|".join(sorted(all_answers))
            
            # Créer le hash
            return hashlib.sha256(combined.encode()).hexdigest()
        
        except Exception as e:
            print(f"❌ Erreur calcul hash : {e}")
            return None
    
    def init_ui(self):
        """Initialise l'interface"""
        layout = QVBoxLayout()
        
        title = QLabel(f"Analyse de Fraude - Worksheet {self.worksheet_num}")
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title.setFont(title_font)
        layout.addWidget(title)
        
        # Label pour le résumé (sera rempli dans analyze_cheating)
        self.summary_label = QLabel()
        self.summary_label.setStyleSheet("font-size: 11px; padding: 5px;")
        layout.addWidget(self.summary_label)
        
        # Tableau des résultats
        self.results_table = QTableWidget()
        self.results_table.setColumnCount(3)
        self.results_table.setHorizontalHeaderLabels([
            "🚨 Hashes Identiques",
            "Nombre d'élèves",
            "Noms des élèves"
        ])
        self.results_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.results_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.results_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        
        layout.addWidget(self.results_table)
        
        # Bouton fermer
        close_btn = QPushButton("Fermer")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn)
        
        self.setLayout(layout)
    
    def analyze_cheating(self):
        """Analyse les réponses pour détecter les doublons"""
        hash_to_students = {}  # {hash: [student_names]}
        
        # Parcourir tous les élèves
        for student_name, worksheets in self.students_dict.items():
            pdf_path = worksheets.get(self.worksheet_num)
            if not pdf_path:
                continue
            
            # Calculer le hash des réponses
            answers_hash = self._calculate_answers_hash(pdf_path)
            if answers_hash:
                if answers_hash not in hash_to_students:
                    hash_to_students[answers_hash] = []
                hash_to_students[answers_hash].append(student_name)
        
        # Trouver les hashes avec plusieurs élèves (fraude probable)
        frauds = {h: students for h, students in hash_to_students.items() if len(students) > 1}
        
        # Calculer le résumé
        total_items = len(hash_to_students)
        suspect_items = len(frauds)
        suspect_students = sum(len(students) for students in frauds.values())
        
        if not frauds:
            # Aucune fraude détectée
            summary_text = f"✅ Aucune fraude détectée - Tous les {total_items} élève(s) ont des réponses uniques"
            self.summary_label.setText(summary_text)
            self.summary_label.setStyleSheet("font-size: 11px; padding: 5px; background-color: #E8F5E9; border-radius: 3px;")
            
            self.results_table.setRowCount(1)
            msg_item = QTableWidgetItem("✅ Aucune fraude détectée - Toutes les réponses sont uniques")
            msg_item.setBackground(QColor(200, 255, 200))
            self.results_table.setItem(0, 0, msg_item)
            return
        
        # Afficher le résumé des fraudes
        summary_text = f"🚨 FRAUDE GRAVÉE : {suspect_items} groupe(s) détecté(s) - {suspect_students} élève(s) suspect(s)"
        self.summary_label.setText(summary_text)
        self.summary_label.setStyleSheet("font-size: 11px; padding: 5px; background-color: #FFEBEE; border-radius: 3px;")
        
        # Afficher les fraudes détectées
        self.results_table.setRowCount(len(frauds))
        
        for row, (answers_hash, students) in enumerate(sorted(frauds.items(), key=lambda x: len(x[1]), reverse=True)):
            students_sorted = sorted(students)
            
            # Hash
            hash_item = QTableWidgetItem(answers_hash[:16] + "...")
            hash_item.setBackground(QColor(255, 200, 200))  # Rouge
            hash_item.setFont(QFont("Courier", 9))
            self.results_table.setItem(row, 0, hash_item)
            
            # Nombre d'élèves
            count_item = QTableWidgetItem(str(len(students)))
            count_item.setBackground(QColor(255, 150, 150))
            count_item.setFont(QFont())
            count_item.font().setBold(True)
            self.results_table.setItem(row, 1, count_item)
            
            # Noms des élèves
            students_names = "\n".join(students_sorted)
            names_item = QTableWidgetItem(students_names)
            names_item.setBackground(QColor(255, 200, 200))
            self.results_table.setItem(row, 2, names_item)
        
        # Redimensionner les lignes
        self.results_table.resizeRowsToContents()


class CorrectionApp(QMainWindow):
    WORKSHEETS = [1, 2, 3, 4]  # Worksheets à corriger
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Correcteur de Formulaires PDF - Renewable Energy")
        self.setGeometry(100, 100, 1600, 900)
        
        # Configuration
        self.config_manager = ConfigManager()
        self.work_dir = self.config_manager.config.get("work_dir") or None
        self.correction_dir = self.config_manager.config.get("correction_dir") or None
        self.students = {}
        self.student_scores = {}  # Scores: {worksheet: {student: pour centage}}
        self.duplicate_hashes = {}  # Hashes dupliqués: {worksheet: {hash: [students]}}
        self.current_worksheet = None
        self.deadline_labels = {}  # Dictionnaire vide pour éviter les erreurs
        
        self.init_ui()
        
        # Charger les répertoires s'ils sont sauvegardés
        if self.work_dir:
            self.work_dir_label.setText(self.work_dir)
            self.work_dir_label.setStyleSheet("color: green; font-weight: bold;")
            self.load_students()
        
        if self.correction_dir:
            self.correction_dir_label.setText(self.correction_dir)
            self.correction_dir_label.setStyleSheet("color: green; font-weight: bold;")
    
    def init_ui(self):
        """Initialise l'interface"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout()
        
        # ===== SECTION CONFIGURATION =====
        config_layout = QHBoxLayout()
        
        # Sélection du répertoire de travail
        config_layout.addWidget(QLabel("Répertoire de travail :"))
        self.work_dir_label = QLabel("Aucun répertoire sélectionné")
        self.work_dir_label.setStyleSheet("color: gray;")
        config_layout.addWidget(self.work_dir_label)
        
        btn_select_work = QPushButton("📁 Sélectionner dossier élèves")
        btn_select_work.clicked.connect(self.select_work_directory)
        config_layout.addWidget(btn_select_work)
        
        # Sélection du répertoire de corrections
        config_layout.addWidget(QLabel("Répertoire corrections :"))
        self.correction_dir_label = QLabel("Aucun répertoire sélectionné")
        self.correction_dir_label.setStyleSheet("color: gray;")
        config_layout.addWidget(self.correction_dir_label)
        
        btn_select_correction = QPushButton("📁 Sélectionner dossier corrections")
        btn_select_correction.clicked.connect(self.select_correction_directory)
        config_layout.addWidget(btn_select_correction)
        
        # Bouton pour importer les dates des deadlines depuis fichier ODS/Excel
        btn_import_dates = QPushButton("📊 Importer dates (ODS/Excel)")
        btn_import_dates.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")
        btn_import_dates.clicked.connect(self.import_deadlines_from_spreadsheet)
        config_layout.addWidget(btn_import_dates)
        
        main_layout.addLayout(config_layout)
        
        # ===== SECTION SÉLECTION WORKSHEET =====
        worksheet_selection_layout = QHBoxLayout()
        label = QLabel("Choisir l'activité :")
        label.setStyleSheet("font-weight: bold; font-size: 14px;")
        worksheet_selection_layout.addWidget(label)
        
        self.worksheet_buttons = {}
        for ws_num in self.WORKSHEETS:
            btn = QPushButton(f"Worksheet {ws_num}")
            btn.setMaximumWidth(120)
            btn.setCheckable(True)
            btn.clicked.connect(lambda checked, n=ws_num: self.on_worksheet_selected(n))
            self.worksheet_buttons[ws_num] = btn
            worksheet_selection_layout.addWidget(btn)
        
        worksheet_selection_layout.addStretch()
        
        # Bouton Correction Batch
        btn_batch_correction = QPushButton("⚡ Correction Batch")
        btn_batch_correction.setMaximumWidth(150)
        btn_batch_correction.setStyleSheet("background-color: #FF6B35; color: white; font-weight: bold;")
        btn_batch_correction.clicked.connect(self.open_batch_correction)
        worksheet_selection_layout.addWidget(btn_batch_correction)
        
        # Bouton Détection de Fraude
        btn_cheat_detection = QPushButton("🔍 Détecter Fraudes")
        btn_cheat_detection.setMaximumWidth(150)
        btn_cheat_detection.setStyleSheet("background-color: #FF4444; color: white; font-weight: bold;")
        btn_cheat_detection.clicked.connect(self.open_cheat_detection)
        worksheet_selection_layout.addWidget(btn_cheat_detection)
        
        main_layout.addLayout(worksheet_selection_layout)
        
        # ===== SECTION PRINCIPALE : SPLITTER =====
        splitter = QSplitter(Qt.Horizontal)
        
        # Colonne gauche : Liste des élèves
        left_widget = QWidget()
        left_layout = QVBoxLayout()
        
        left_layout.addWidget(QLabel("Élèves :"))
        self.student_list = QListWidget()
        self.student_list.setMinimumWidth(200)
        self.student_list.itemClicked.connect(self.on_student_selected)
        left_layout.addWidget(self.student_list)
        
        # Label pour afficher le nombre d'élèves
        self.student_count_label = QLabel("")
        self.student_count_label.setStyleSheet("color: #333; font-size: 14px; font-weight: bold; padding: 12px; background-color: #f5f5f5; border-radius: 4px; border: 1px solid #ddd;")
        left_layout.addWidget(self.student_count_label)
        
        left_widget.setLayout(left_layout)
        splitter.addWidget(left_widget)
        
        # Colonne droite : Affichage des PDFs et informations
        right_widget = QWidget()
        right_layout = QVBoxLayout()
        
        # Tableau d'informations sur le fichier
        self.info_table = QTableWidget(5, 2)
        self.info_table.setHorizontalHeaderLabels(["Propriété", "Valeur"])
        self.info_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.info_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.info_table.setMaximumHeight(300)
        right_layout.addWidget(self.info_table)
        
        # Affichage des PDFs côte à côte
        pdf_layout = QHBoxLayout()
        
        # Formulaire étudiant
        student_pdf_layout = QVBoxLayout()
        student_pdf_layout.addWidget(QLabel("Travail étudiant :"))
        self.student_pdf_label = ClickableLabel("Aucun fichier")
        self.student_pdf_label.setAlignment(Qt.AlignCenter)
        self.student_pdf_label.setStyleSheet("border: 1px solid gray; background-color: #f0f0f0; cursor: pointer;")
        self.student_pdf_label.setMinimumHeight(500)
        student_pdf_layout.addWidget(self.student_pdf_label)
        
        pdf_layout.addLayout(student_pdf_layout)
        
        # Bouton Correction entre les deux PDFs
        correction_button_layout = QVBoxLayout()
        correction_button_layout.addStretch()
        
        btn_correction = QPushButton("📋\nCORRECTION\nCOMPARATIVE")
        btn_correction.setMaximumWidth(100)
        btn_correction.setMinimumHeight(100)
        btn_correction.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                font-weight: bold;
                border-radius: 5px;
                border: 2px solid #F57C00;
            }
            QPushButton:hover {
                background-color: #F57C00;
            }
            QPushButton:pressed {
                background-color: #E65100;
            }
        """)
        btn_correction.clicked.connect(self.open_correction_interface)
        correction_button_layout.addWidget(btn_correction)
        
        correction_button_layout.addStretch()
        
        pdf_layout.addLayout(correction_button_layout)
        
        # Correction
        correction_pdf_layout = QVBoxLayout()
        correction_pdf_layout.addWidget(QLabel("Correction :"))
        self.correction_pdf_label = ClickableLabel("Aucun fichier")
        self.correction_pdf_label.setAlignment(Qt.AlignCenter)
        self.correction_pdf_label.setStyleSheet("border: 1px solid gray; background-color: #f0f0f0; cursor: pointer;")
        self.correction_pdf_label.setMinimumHeight(500)
        correction_pdf_layout.addWidget(self.correction_pdf_label)
        
        pdf_layout.addLayout(correction_pdf_layout)
        
        right_layout.addLayout(pdf_layout)
        
        right_widget.setLayout(right_layout)
        splitter.addWidget(right_widget)
        
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 2)
        
        main_layout.addWidget(splitter)
        
        central_widget.setLayout(main_layout)
        
        # Initialiser l'interface
        self.current_worksheet = None
        self.update_student_list()
    
    def select_work_directory(self):
        """Sélectionne le répertoire de travail des élèves"""
        directory = QFileDialog.getExistingDirectory(
            self, "Sélectionner le répertoire de travail des élèves"
        )
        if directory:
            self.work_dir = directory
            self.config_manager.set_work_dir(directory)
            self.work_dir_label.setText(directory)
            self.work_dir_label.setStyleSheet("color: green; font-weight: bold;")
            self.load_students()
    
    def select_correction_directory(self):
        """Sélectionne le répertoire des corrections"""
        directory = QFileDialog.getExistingDirectory(
            self, "Sélectionner le répertoire des corrections"
        )
        if directory:
            self.correction_dir = directory
            self.config_manager.set_correction_dir(directory)
            self.correction_dir_label.setText(directory)
            self.correction_dir_label.setStyleSheet("color: green; font-weight: bold;")
    
    def _format_deadline(self, ws_num):
        """Formate l'affichage de la deadline"""
        deadline = self.config_manager.get_deadline(ws_num)
        if deadline:
            return deadline.strftime("%d/%m/%Y")
        return "Non définie"
    
    def set_worksheet_deadline(self, ws_num):
        """Définit la deadline pour un worksheet spécifique"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Définir deadline - Worksheet {ws_num}")
        layout = QVBoxLayout()
        
        layout.addWidget(QLabel(f"Choisir la deadline pour Worksheet {ws_num} :"))
        
        calendar = QCalendarWidget()
        current_deadline = self.config_manager.get_deadline(ws_num)
        if current_deadline:
            calendar.setSelectedDate(QDate(current_deadline.year, current_deadline.month, current_deadline.day))
        else:
            calendar.setSelectedDate(QDate.currentDate())
        layout.addWidget(calendar)
        
        btn_layout = QHBoxLayout()
        ok_button = QPushButton("Valider")
        cancel_button = QPushButton("Annuler")
        
        ok_button.clicked.connect(lambda: self.confirm_worksheet_deadline(ws_num, calendar.selectedDate(), dialog))
        cancel_button.clicked.connect(dialog.reject)
        
        btn_layout.addWidget(ok_button)
        btn_layout.addWidget(cancel_button)
        layout.addLayout(btn_layout)
        
        dialog.setLayout(layout)
        dialog.exec_()
    
    def confirm_worksheet_deadline(self, ws_num, q_date, dialog):
        """Confirme la deadline sélectionnée"""
        deadline = q_date.toPyDate()
        self.config_manager.set_deadline(ws_num, deadline)
        # Mettre à jour le label seulement s'il existe
        if ws_num in self.deadline_labels:
            self.deadline_labels[ws_num].setText(deadline.strftime("%d/%m/%Y"))
        dialog.close()
    
    def import_deadlines_from_spreadsheet(self):
        """Importe les deadlines depuis un fichier Calc/Excel"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Sélectionner le fichier de planning",
            str(Path.home()),
            "Spreadsheet Files (*.xlsx *.ods *.calc);;Excel Files (*.xlsx);;ODS Files (*.ods);;All Files (*)"
        )
        
        if not file_path:
            return
        
        try:
            print(f"📂 Chargement des dates depuis: {file_path}")
            # Sauvegarder le chemin du fichier pour les exports futurs
            self.config_manager.config["last_spreadsheet_path"] = str(file_path)
            self.config_manager.save_config()
            
            result = SpreadsheetParser.parse_spreadsheet(file_path)
            
            if not result:
                QMessageBox.warning(self, "Aucune date trouvée", 
                    "Aucune deadline n'a pu être extraite du fichier.")
                return
            
            # Vérifier si c'est un dictionnaire par élève ou global
            is_per_student = isinstance(result, dict) and any(
                isinstance(v, dict) for v in result.values()
            )
            
            if is_per_student:
                # Format par élève (nouveau format de présence)
                self._apply_student_deadlines(result)
            else:
                # Format global (ancien format)
                self._apply_global_deadlines(result)
        
        except ImportError as e:
            QMessageBox.critical(self, "Erreur de dépendance", 
                f"Les outils nécessaires ne sont pas installés:\n{str(e)}")
        except Exception as e:
            QMessageBox.critical(self, "Erreur lors de l'importation", 
                f"Une erreur s'est produite:\n{str(e)}")
            print(f"❌ Erreur: {e}")
            import traceback
            traceback.print_exc()
    
    def _apply_global_deadlines(self, deadlines):
        """Applique les deadlines globales (same date pour tous les élèves)"""
        dates_applied = []
        for ws_num, deadline_date in deadlines.items():
            if isinstance(ws_num, int) and 1 <= ws_num <= 4:
                self.config_manager.set_deadline(ws_num, deadline_date)
                # Mettre à jour le label seulement s'il existe
                if ws_num in self.deadline_labels:
                    self.deadline_labels[ws_num].setText(deadline_date.strftime("%d/%m/%Y"))
                dates_applied.append(f"Worksheet {ws_num}: {deadline_date.strftime('%d/%m/%Y')}")
        
        if dates_applied:
            QMessageBox.information(self, "✓ Deadlines importées avec succès", 
                "\n".join(dates_applied))
            print(f"✅ {len(dates_applied)} deadline(s) chargée(s)")
        else:
            QMessageBox.warning(self, "Aucune deadline trouvée", 
                "Aucune deadline valide n'a pu être extraite du fichier.")
    
    def _apply_student_deadlines(self, student_deadlines):
        """Applique les deadlines par élève depuis le fichier de présence"""
        total_dates = 0
        students_updated = set()
        
        # Fonction pour matcher les noms du tableur avec les dossiers  
        def find_matching_folder(tableur_name, folder_names):
            """
            Essaie de trouver le dossier correspondant au nom du tableur.
            Convention: dossier = nom_de_famille + [première_lettre_prénom] + [chiffre]
            Ex: ABATE Tom → ABATET, AZAUBERT Esteban → AZAUBERT1
            """
            # Chercher une correspondance exacte (case-insensitive, sans espaces)
            tableur_normalized = tableur_name.lower().replace(" ", "")
            for folder in folder_names:
                if folder.lower() == tableur_normalized:
                    return folder
            
            # Extraire le nom de famille et le prénom
            parts = tableur_name.split()
            if len(parts) < 1:
                return None
            
            family_name = parts[0].lower()  # ex: "abate" de "ABATE Tom"
            first_initial = parts[1][0].lower() if len(parts) > 1 else ""  # ex: "t" de "Tom"
            
            # Chercher un dossier qui commence par le nom de famille
            candidates = []
            for folder in folder_names:
                folder_lower = folder.lower()
                
                # Le dossier doit commencer par le nom de famille
                if folder_lower.startswith(family_name):
                    remainder = folder_lower[len(family_name):]
                    
                    # Vérifier que ce qui suit est la première lettre du prénom ou un chiffre
                    # ou rien du tout
                    if (not remainder or  # Correspondance exacte du nom de famille
                        remainder[0] == first_initial or  # Commence par la première lettre du prénom
                        remainder[0].isdigit()):  # Commence par un chiffre (doublon)
                        candidates.append(folder)
            
            # Retourner le candidat trouvé (préférer le plus court = moins d'ajouts)
            if candidates:
                return min(candidates, key=len)
            
            return None
        
        print(f"\n🔍 Matching des noms (tableur → dossier):")
        
        for student_name, ws_deadlines in student_deadlines.items():
            if not ws_deadlines:
                continue
            
            # Trouver le dossier correspondant
            actual_student_name = find_matching_folder(student_name, list(self.students.keys()))
            
            if actual_student_name:
                if actual_student_name != student_name:
                    print(f"  ✓ '{student_name}' → '{actual_student_name}'")
                
                # Sauvegarder les deadlines pour cet élève
                for ws_num, deadline_date in ws_deadlines.items():
                    if isinstance(ws_num, int) and 1 <= ws_num <= 4:
                        self._save_student_deadline(actual_student_name, ws_num, deadline_date)
                        total_dates += 1
                        students_updated.add(actual_student_name)
            else:
                print(f"  ❌ '{student_name}' → AUCUN MATCH")
        
        if students_updated:
            message = f"✓ Deadlines importées pour {len(students_updated)} élève(s)\n"
            message += f"Total: {total_dates} deadline(s)\n\n"
            message += "Exemples:\n"
            for i, (student, deadlines) in enumerate(student_deadlines.items()):
                if i >= 3:  # Montrer les 3 premiers
                    message += f"... et {len(students_updated)-3} autre(s)"
                    break
                if deadlines:
                    message += f"• {student}: WS1→{list(deadlines.values())[0] if 1 in deadlines else '?'}\n"
            
            QMessageBox.information(self, "✓ Deadlines importées avec succès", message)
            print(f"✅ {total_dates} deadline(s) chargée(s) pour {len(students_updated)} élève(s)")
        else:
            QMessageBox.warning(self, "Aucune deadline trouvée", 
                "Aucune deadline valide n'a pu être extraite du fichier.")
    
    def _save_student_deadline(self, student_name, ws_num, deadline_date):
        """Sauvegarde une deadline pour un élève spécifique"""
        if "student_deadlines" not in self.config_manager.config:
            self.config_manager.config["student_deadlines"] = {}
        
        if student_name not in self.config_manager.config["student_deadlines"]:
            self.config_manager.config["student_deadlines"][student_name] = {}
        
        self.config_manager.config["student_deadlines"][student_name][f"worksheet{ws_num}"] = str(deadline_date)
        self.config_manager.save_config()
    
    def _get_student_deadline(self, student_name, ws_num):
        """Récupère la deadline pour un élève spécifique"""
        if "student_deadlines" not in self.config_manager.config:
            return None
        
        if student_name not in self.config_manager.config["student_deadlines"]:
            return None
        
        deadline_str = self.config_manager.config["student_deadlines"][student_name].get(
            f"worksheet{ws_num}"
        )
        
        if deadline_str:
            try:
                result = datetime.strptime(deadline_str, "%Y-%m-%d").date()
                return result
            except Exception as e:
                print(f"❌ Erreur parsing deadline {student_name} WS{ws_num}: {deadline_str} -> {e}")
                return None
        
        return None
    
    def load_students(self):
        """Charge la liste complète des élèves et leurs worksheets"""
        if not self.work_dir:
            print("❌ Aucun répertoire de travail défini")
            return
        
        self.students = {}
        students_missing_files = []  # Élèves sans fichiers détectés
        
        work_path = Path(self.work_dir)
        
        if not work_path.exists():
            print(f"❌ Le répertoire {work_path} n'existe pas")
            return
        
        print(f"📂 Chargement depuis : {work_path}")
        
        # Parcourir les dossiers d'élèves
        try:
            for student_folder in sorted(work_path.iterdir()):
                if student_folder.is_dir():
                    student_name = student_folder.name
                    worksheets = {}
                    
                    # Chercher les fichiers worksheets avec la nouvelle méthode flexible
                    for ws_num in self.WORKSHEETS:
                        # D'abord essayer la détection automatique
                        worksheet_file = ConfigManager.find_worksheet_file(student_folder, ws_num)
                        
                        # Sinon, chercher une sélection manuelle sauvegardée
                        if not worksheet_file:
                            worksheet_file = self.config_manager.get_manual_selection(student_name, ws_num)
                        
                        if worksheet_file:
                            worksheets[ws_num] = worksheet_file
                            if ConfigManager.find_worksheet_file(student_folder, ws_num):
                                print(f"  ✓ {student_name}: worksheet{ws_num} trouvé -> {worksheet_file.name}")
                            else:
                                print(f"  ✓ {student_name}: worksheet{ws_num} (manuel) -> {worksheet_file.name}")
                    
                    # Ajouter TOUJOURS l'élève, même sans fichiers détectés
                    self.students[student_name] = worksheets
                    
                    # Marquer les élèves incomplets
                    if not worksheets:
                        students_missing_files.append(student_name)
                        print(f"  ⚠️ {student_name}: aucun fichier détecté (sélection manuelle requise)")
        
        except Exception as e:
            print(f"❌ Erreur lors du chargement : {e}")
            import traceback
            traceback.print_exc()
        
        print(f"✅ {len(self.students)} élève(s) chargé(s)")
        for name, worksheets in self.students.items():
            print(f"  - {name}: {list(worksheets.keys())}")
        
        # Si des élèves n'ont pas de fichiers, proposer sélection manuelle
        if students_missing_files:
            print(f"⚠️ {len(students_missing_files)} élève(s) nécessitent une sélection manuelle des fichiers")
            for name in students_missing_files:
                print(f"   - {name}")
    
    def calculate_all_scores(self, ws_num):
        """Calcule les scores de tous les élèves pour un worksheet donné"""
        if not self.correction_dir:
            print(f"⚠️ Pas de répertoire de correction défini")
            return {}
        
        # Importer les fonctions de comparaison depuis correction.py
        try:
            from correction import extract_fields_from_pdf, compare_pdf_with_reference
        except ImportError:
            print("⚠️ Impossible d'importer les fonctions de comparaison")
            return {}
        
        correction_path = Path(self.correction_dir) / f"correction_worksheet{ws_num}.pdf"
        if not correction_path.exists():
            print(f"⚠️ Fichier de correction non trouvé: {correction_path}")
            return {}
        
        # Extraire les champs de référence de la correction
        try:
            reference_fields = extract_fields_from_pdf(str(correction_path))
            if not reference_fields:
                print(f"⚠️ Aucun champ détecté dans la correction")
                return {}
        except Exception as e:
            print(f"⚠️ Erreur lecture correction: {e}")
            import traceback
            traceback.print_exc()
            return {}
        
        scores = {}
        
        # Calculer le score pour chaque élève
        for student_name, worksheets in self.students.items():
            if ws_num in worksheets:
                student_file = worksheets[ws_num]
                if isinstance(student_file, str):
                    student_file = Path(student_file)
                
                if student_file.exists():
                    try:
                        result = compare_pdf_with_reference(
                            str(student_file), 
                            reference_fields
                        )
                        # Extraire les valeurs de retour (peut être tuple ou liste)
                        if isinstance(result, (list, tuple)) and len(result) >= 2:
                            score = result[0]
                            total = result[1]
                        else:
                            score, total = 0, 1
                        
                        percentage = int((score / total * 100)) if total > 0 else 0
                        scores[student_name] = percentage
                    except Exception as e:
                        print(f"⚠️ Erreur calcul score {student_name}: {e}")
                        scores[student_name] = 0
        
        # Sauvegarder les scores
        self.student_scores[ws_num] = scores
        print(f"✅ Scores calculés pour WS{ws_num}: {len(scores)} élèves")
        return scores
    
    def detect_duplicate_hashes(self, ws_num):
        """Détecte les fichiers avec le même hash (copies)"""
        hashes = {}
        
        for student_name, worksheets in self.students.items():
            if ws_num in worksheets:
                student_file = worksheets[ws_num]
                if isinstance(student_file, str):
                    student_file = Path(student_file)
                
                if student_file.exists():
                    try:
                        # Calculer le hash du fichier
                        sha256_hash = hashlib.sha256()
                        with open(student_file, "rb") as f:
                            for byte_block in iter(lambda: f.read(4096), b""):
                                sha256_hash.update(byte_block)
                        file_hash = sha256_hash.hexdigest()
                        
                        # Ajouter à la liste des hashes
                        if file_hash not in hashes:
                            hashes[file_hash] = []
                        hashes[file_hash].append(student_name)
                    except Exception as e:
                        print(f"⚠️ Erreur calcul hash {student_name}: {e}")
        
        # Garder seulement les hashes dupliqués
        duplicates = {hash_val: students for hash_val, students in hashes.items() if len(students) > 1}
        
        # Inverser pour avoir student -> bool
        student_has_duplicate = {}
        for students in duplicates.values():
            for student_name in students:
                student_has_duplicate[student_name] = True
        
        self.duplicate_hashes[ws_num] = student_has_duplicate
        return student_has_duplicate
    
    def update_student_list(self):
        """Met à jour la liste des élèves en fonction du worksheet sélectionné"""
        self.student_list.clear()
        
        # Si aucun worksheet sélectionné, afficher tous les élèves
        if not self.current_worksheet:
            self.student_list.addItem("Sélectionnez un worksheet d'abord")
            # Afficher le nombre total d'élèves même sans worksheet sélectionné
            total_students = len(self.students)
            self.student_count_label.setText(f"{total_students} élève(s) détecté(s)")
            return
        
        # Afficher les élèves avec plusieurs catégories
        has_worksheet_auto = []
        has_worksheet_manual = []
        missing_worksheet = []
        
        for student_name in sorted(self.students.keys()):
            if self.current_worksheet in self.students[student_name]:
                # Vérifier si c'est une détection automatique ou manuelle
                is_auto = bool(ConfigManager.find_worksheet_file(
                    Path(self.work_dir) / student_name, 
                    self.current_worksheet
                ))
                if is_auto:
                    has_worksheet_auto.append(student_name)
                else:
                    has_worksheet_manual.append(student_name)
            else:
                missing_worksheet.append(student_name)
        
        # Fonction helper pour formater le nom avec deadline et score
        def format_name_with_score(student_name):
            """Formate le nom avec la deadline, score et badge si duplicat"""
            deadline = self._get_student_deadline(student_name, self.current_worksheet)
            if not deadline:
                deadline = self.config_manager.get_deadline(self.current_worksheet)
            
            # Ajouter le score s'il existe
            score_text = ""
            if self.current_worksheet in self.student_scores:
                if student_name in self.student_scores[self.current_worksheet]:
                    score = self.student_scores[self.current_worksheet][student_name]
                    score_text = f" [{score}%]"
            
            # Ajouter le badge si hash dupliqué
            duplicate_badge = ""
            if self.current_worksheet in self.duplicate_hashes:
                if self.duplicate_hashes[self.current_worksheet].get(student_name, False):
                    duplicate_badge = " 🔴"  # Badge pour copie détectée
            
            if deadline:
                return f"{student_name}{score_text}{duplicate_badge} (Deadline: {deadline.strftime('%d/%m/%Y')})"
            return f"{student_name}{score_text}{duplicate_badge}"
        
        # Ajouter les élèves avec détection automatique (vert)
        for student_name in has_worksheet_auto:
            item = QListWidgetItem(f"✓ {format_name_with_score(student_name)}")
            item.setData(Qt.UserRole, student_name)
            item.setForeground(QColor(0, 128, 0))  # Vert
            self.student_list.addItem(item)
        
        # Ajouter les élèves avec fichier manuellement sélectionné (bleu)
        if has_worksheet_manual:
            separator = QListWidgetItem("--- Fichier sélectionné manuellement ---")
            separator.setFlags(separator.flags() & ~Qt.ItemIsSelectable)
            separator.setForeground(QColor(128, 128, 128))
            self.student_list.addItem(separator)
            
            for student_name in has_worksheet_manual:
                item = QListWidgetItem(f"📋 {format_name_with_score(student_name)}")
                item.setData(Qt.UserRole, student_name)
                item.setForeground(QColor(0, 0, 255))  # Bleu
                self.student_list.addItem(item)
        
        # Ajouter les élèves qui n'ont pas le worksheet (avec option de sélection manuelle)
        if missing_worksheet:
            separator = QListWidgetItem("--- Sans fichier détecté (sélection manuelle) ---")
            separator.setFlags(separator.flags() & ~Qt.ItemIsSelectable)
            separator.setForeground(QColor(128, 128, 128))
            self.student_list.addItem(separator)
            
            for student_name in missing_worksheet:
                item = QListWidgetItem(f"⚠️ {format_name_with_score(student_name)}")
                item.setData(Qt.UserRole, student_name)
                item.setForeground(QColor(255, 0, 0))  # Rouge
                self.student_list.addItem(item)
        
        # Mettre à jour le label de décompte
        total_students = len(self.students)
        count_text = f"{total_students} élève(s) détecté(s)"
        if self.current_worksheet:
            count_text += f" • {len(has_worksheet_auto)} ✓ • {len(has_worksheet_manual)} 📋 • {len(missing_worksheet)} ⚠️"
        self.student_count_label.setText(count_text)
    
    def on_worksheet_selected(self, ws_num):
        """Gère la sélection d'un worksheet"""
        # Décocher tous les boutons
        for btn in self.worksheet_buttons.values():
            btn.setChecked(False)
        
        # Cocher le bouton sélectionné
        self.worksheet_buttons[ws_num].setChecked(True)
        
        self.current_worksheet = ws_num
        
        # Calculer les scores pour tous les élèves
        self.calculate_all_scores(ws_num)
        
        # Détecter les doublons (hashes identiques)
        self.detect_duplicate_hashes(ws_num)
        
        self.update_student_list()
        
        # Effacer l'affichage
        self.student_pdf_label.setText("Sélectionnez un élève")
        self.correction_pdf_label.setText("---")
        self.info_table.setRowCount(0)
        
        print(f"📝 Worksheet {ws_num} sélectionné")
    
    def on_student_selected(self, item):
        """Gère la sélection d'un élève"""
        if not self.current_worksheet:
            return
        
        student_name = item.data(Qt.UserRole)
        if not student_name or student_name not in self.students:
            return
        
        # Récupérer le fichier du worksheet sélectionné
        ws_path = self.students[student_name].get(self.current_worksheet)
        
        # Si le fichier n'est pas détecté, proposer une sélection manuelle
        if not ws_path:
            self.select_file_manually(student_name)
            return
        
        # Afficher le fichier détecté
        self.display_student_work(student_name, ws_path)
    
    def select_file_manually(self, student_name):
        """Permet de sélectionner manuellement le fichier d'un élève"""
        student_folder = Path(self.work_dir) / student_name
        
        # Si le dossier standard n'existe pas, chercher un dossier alternatif
        start_folder = str(student_folder) if student_folder.exists() else str(Path(self.work_dir))
        
        # Créer un message informatif
        msg = f"Sélectionner le fichier worksheet {self.current_worksheet} pour {student_name}"
        if not student_folder.exists():
            msg += f"\n\n⚠️ Le dossier {student_folder} n'existe pas.\nLes fichiers peuvent être ailleurs."
        
        # Ouvrir un sélecteur de fichier
        file_path, _ = QFileDialog.getOpenFileName(
            self, 
            msg,
            start_folder,
            "PDF Files (*.pdf);;All Files (*)"
        )
        
        if file_path:
            file_path = Path(file_path)
            
            # Sauvegarder le choix manuel dans les données de l'élève
            self.students[student_name][self.current_worksheet] = file_path
            
            # Sauvegarder dans la configuration
            self.config_manager.save_manual_selection(student_name, self.current_worksheet, file_path)
            
            # Afficher le fichier sélectionné
            self.display_student_work(student_name, file_path)
            
            print(f"✅ {student_name}: worksheet {self.current_worksheet} assigné manuellement -> {file_path.name}")
    
    def display_student_work(self, student_name, ws_path):
        """Affiche le travail d'un élève et la correction correspondante"""
        ws_path = Path(ws_path)
        
        # Mettre à jour les informations avec le numéro du worksheet
        self.update_file_info(ws_path, student_name, self.current_worksheet)
        
        # Afficher le PDF de l'étudiant
        self.display_pdf(ws_path, self.student_pdf_label)
        
        # Afficher la correction correspondante
        if self.correction_dir:
            correction_path = Path(self.correction_dir) / f"correction_worksheet{self.current_worksheet}.pdf"
            if correction_path.exists():
                self.display_pdf(correction_path, self.correction_pdf_label)
            else:
                self.correction_pdf_label.setText(f"❌ Correction non trouvée\n{correction_path.name}")

    
    def update_file_info(self, file_path, student_name, ws_num=None):
        """Met à jour le tableau d'informations avec détails deadline"""
        file_path = Path(file_path)
        
        # Information du fichier
        stat = file_path.stat()
        file_date = datetime.fromtimestamp(stat.st_mtime)
        file_size = stat.st_size / 1024  # en KB
        
        # Préparation des données
        data = [
            ("Étudiant", student_name),
            ("Fichier", file_path.name),
            ("Date de modification", file_date.strftime("%d/%m/%Y %H:%M:%S")),
            ("Taille", f"{file_size:.1f} KB"),
        ]
        
        # Ajouter les deadlines pour TOUS les worksheets
        deadline_colors = {}  # Stocke les couleurs pour chaque ligne deadline
        
        for ws_number in self.WORKSHEETS:
            # D'abord chercher la deadline spécifique de l'élève
            deadline = self._get_student_deadline(student_name, ws_number)
            
            # Sinon, utiliser la deadline globale
            if not deadline:
                deadline = self.config_manager.get_deadline(ws_number)
            
            # Déterminer le statut
            deadline_status = "❓ Non définie"
            color = QColor(150, 150, 150)  # Gris
            
            if deadline:
                if ws_number == ws_num:
                    # Pour le worksheet courant, vérifier si en retard
                    if file_date.date() > deadline:
                        deadline_status = f"⚠️ RETARD ({deadline.strftime('%d/%m/%Y')})"
                        color = QColor(255, 0, 0)  # Rouge
                    else:
                        deadline_status = f"✓ À temps ({deadline.strftime('%d/%m/%Y')})"
                        color = QColor(0, 128, 0)  # Vert
                else:
                    # Pour les autres worksheets, juste afficher la date
                    deadline_status = deadline.strftime('%d/%m/%Y')
                    color = QColor(0, 0, 255)  # Bleu
            
            deadline_colors[len(data)] = color
            data.append((f"Deadline WS{ws_number}", deadline_status))
        
        # Remplir le tableau
        self.info_table.setRowCount(len(data))
        for row, (key, value) in enumerate(data):
            item_key = QTableWidgetItem(key)
            item_value = QTableWidgetItem(value)
            
            # Colorer les lignes de deadline
            if row in deadline_colors:
                item_value.setForeground(deadline_colors[row])
                item_value.setFont(QFont("Arial", 9, QFont.Bold))
            
            self.info_table.setItem(row, 0, item_key)
            self.info_table.setItem(row, 1, item_value)
    
    def display_pdf(self, pdf_path, label_widget):
        """Affiche un aperçu du PDF"""
        try:
            pdf_path = Path(pdf_path)
            if not pdf_path.exists():
                label_widget.setText("❌ Fichier non trouvé")
                return
            
            # Définir le chemin pour la vignette cliquable
            if isinstance(label_widget, ClickableLabel):
                label_widget.set_pdf_path(str(pdf_path))
            
            # Convertir la première page en image
            images = convert_from_path(str(pdf_path), first_page=1, last_page=1, dpi=150)
            if images:
                img = images[0]
                
                # Redimensionner pour adapter à la taille du label
                max_width = 350
                max_height = 500
                img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
                
                # Convertir PIL Image en QPixmap via données brutes
                with io.BytesIO() as output:
                    img.save(output, format="PNG")
                    data = output.getvalue()
                
                pixmap = QPixmap()
                pixmap.loadFromData(data, "PNG")
                
                label_widget.setPixmap(pixmap)
                label_widget.setAlignment(Qt.AlignCenter)
            else:
                label_widget.setText("❌ Impossible de convertir le PDF")
        
        except ImportError:
            label_widget.setText("⚠️ pdf2image non installé\nInstallation requise")
        except Exception as e:
            label_widget.setText(f"❌ Erreur : {str(e)}")
    
    def extract_pdf_fields(self, pdf_path):
        """
        Extrait les champs d'un formulaire PDF avec leurs positions
        Essaie d'abord les champs AcroForm, puis détecte les zones de saisie
        """
        fields = {}
        
        try:
            # Méthode 1: Essayer d'extraire les champs AcroForm (formulaires interactifs)
            if PYPDF2_AVAILABLE:
                try:
                    reader = PdfReader(str(pdf_path))
                    if reader.get_fields():
                        field_names = list(reader.get_fields().keys())
                        for i, field_name in enumerate(field_names):
                            # Simplifier les noms
                            simple_name = field_name.split('.')[-1] if '.' in field_name else field_name
                            
                            # Essayer de récupérer la position du champ
                            position = {}
                            try:
                                field_obj = reader.get_fields()[field_name]
                                if '/Rect' in field_obj:
                                    rect = field_obj['/Rect']
                                    position = {
                                        'x0': float(rect[0]),
                                        'y0': float(rect[1]),
                                        'x1': float(rect[2]),
                                        'y1': float(rect[3])
                                    }
                            except:
                                pass
                            
                            fields[f"field_{i+1}"] = {
                                'name': simple_name, 
                                'type': 'AcroForm',
                                'original_name': field_name,
                                'position': position
                            }
                        
                        if fields:
                            print(f"✓ {len(fields)} champs AcroForm détectés dans {Path(pdf_path).name}")
                            return fields
                except Exception as e:
                    print(f"⚠️ Erreur lors de la lecture des champs AcroForm: {e}")
            
            # Méthode 2: Détecter les zones de saisie par tableaux et rectangles
            with pdfplumber.open(str(pdf_path)) as pdf:
                if len(pdf.pages) > 0:
                    page = pdf.pages[0]
                    
                    # Chercher les tableaux (zones structurées)
                    tables = page.find_tables()
                    if tables:
                        field_count = 1
                        for table_idx, table in enumerate(tables):
                            for i in range(len(table)):
                                fields[f"field_{field_count}"] = {
                                    'name': f'Tableau {table_idx+1} - Cellule {i+1}',
                                    'type': 'Table',
                                    'position': {
                                        'x0': table['x0'],
                                        'y0': table['top'],
                                        'x1': table['x1'],
                                        'y1': table['bottom']
                                    }
                                }
                                field_count += 1
                    
                    # Chercher les rectangles (boîtes de saisie)
                    if hasattr(page, 'rects') and page.rects:
                        rect_count = 1
                        for rect in page.rects[:10]:  # Limiter à 10 rectangles
                            field_key = f"field_rect_{rect_count}"
                            fields[field_key] = {
                                'name': f'Zone {rect_count}',
                                'type': 'Rectangle',
                                'position': {
                                    'x0': rect['x0'],
                                    'y0': rect['top'],
                                    'x1': rect['x1'],
                                    'y1': rect['bottom']
                                }
                            }
                            rect_count += 1
                    
                    # Si on a trouvé des champs, les retourner
                    if fields:
                        print(f"✓ {len(fields)} zones de saisie détectées dans {Path(pdf_path).name}")
                        return fields
        
        except Exception as e:
            print(f"❌ Erreur extraction champs PDF: {e}")
        
        print(f"⚠️ Aucun champ détecté dans {Path(pdf_path).name}")
        return {}
    
    def setup_notation_fields(self, student_name, worksheet_num):
        """Configure les champs de notation pour l'étudiant/worksheet courant"""
        # Nettoyer les spinbox précédents
        while self.notation_fields_layout.count():
            item = self.notation_fields_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        self.notation_spinboxes = {}
        self.field_positions = {}  # Stocker les positions pour la localisation
        
        # Récupérer le PDF de l'étudiant
        if not student_name or worksheet_num not in self.students.get(student_name, {}):
            self.total_score_label.setText("Aucun fichier sélectionné")
            return
        
        pdf_path = self.students[student_name][worksheet_num]
        fields = self.extract_pdf_fields(pdf_path)
        
        if not fields:
            self.total_score_label.setText("Aucun champ détecté")
            return
        
        # Créer les spinbox pour chaque champ
        for i, (field_key, field_info) in enumerate(fields.items()):
            row = i // 3  # 3 colonnes
            col = i % 3
            
            # Label du champ (afficher le vrai nom)
            label = QLabel(field_info['name'])
            label.setMaximumWidth(150)
            label.setToolTip(f"Champ: {field_info['name']}")
            
            # SpinBox pour la note (pas de maximum, 1 par défaut)
            spinbox = QSpinBox()
            spinbox.setMinimum(0)
            spinbox.setMaximum(999)  # Limite haute au cas où
            spinbox.setValue(1)  # 1 par défaut
            spinbox.setMaximumWidth(60)
            
            # Charger la note sauvegardée si elle existe
            saved_score = self.load_field_score(student_name, worksheet_num, field_key)
            if saved_score is not None:
                spinbox.setValue(saved_score)
            
            # Connecter le signal pour sauvegarder à chaque modification
            spinbox.valueChanged.connect(
                lambda val, sn=student_name, ws=worksheet_num, fk=field_key: 
                self._on_field_score_changed(val, sn, ws, fk)
            )
            
            self.notation_spinboxes[field_key] = spinbox
            self.field_positions[field_key] = field_info.get('position', {})
            
            # Bouton "Voir" pour localiser le champ dans le PDF
            view_btn = QPushButton("👁️")
            view_btn.setMaximumWidth(35)
            view_btn.setToolTip(f"Localiser '{field_info['name']}' dans le PDF")
            view_btn.clicked.connect(
                lambda checked=False, fn=field_info['name']: 
                self._highlight_field_in_pdf(fn)
            )
            
            # Ajouter à la grille
            self.notation_fields_layout.addWidget(label, row, col*3)
            self.notation_fields_layout.addWidget(spinbox, row, col*3+1)
            self.notation_fields_layout.addWidget(view_btn, row, col*3+2)
        
        # Mettre à jour la note totale
        self.update_total_score()
    
    def _on_field_score_changed(self, value, student_name, worksheet_num, field_key):
        """Callback quand la note d'un champ change"""
        self.save_field_score(student_name, worksheet_num, field_key, value)
        self.update_total_score()
    
    def update_total_score(self):
        """Met à jour la note totale"""
        total = sum(spinbox.value() for spinbox in self.notation_spinboxes.values())
        max_score = len(self.notation_spinboxes)
        
        if max_score > 0:
            self.total_score_label.setText(f"Note totale: {total} / {max_score}")
        else:
            self.total_score_label.setText("Note totale: 0 / 0")
    
    def save_field_score(self, student_name, worksheet_num, field_key, score):
        """Sauvegarde la note d'un champ pour un élève"""
        if "student_scores" not in self.config_manager.config:
            self.config_manager.config["student_scores"] = {}
        
        if student_name not in self.config_manager.config["student_scores"]:
            self.config_manager.config["student_scores"][student_name] = {}
        
        ws_key = f"worksheet{worksheet_num}"
        if ws_key not in self.config_manager.config["student_scores"][student_name]:
            self.config_manager.config["student_scores"][student_name][ws_key] = {}
        
        self.config_manager.config["student_scores"][student_name][ws_key][field_key] = score
        self.config_manager.save_config()
    
    def load_field_score(self, student_name, worksheet_num, field_key):
        """Charge la note d'un champ pour un élève"""
        if "student_scores" not in self.config_manager.config:
            return None
        
        if student_name not in self.config_manager.config["student_scores"]:
            return None
        
        ws_key = f"worksheet{worksheet_num}"
        if ws_key not in self.config_manager.config["student_scores"][student_name]:
            return None
        
        field_scores = self.config_manager.config["student_scores"][student_name][ws_key]
        return field_scores.get(field_key)
    
    def _highlight_field_in_pdf(self, field_name):
        """Affiche le champ surligné dans la vignette PDF"""
        # Récupérer le PDF de l'étudiant actuellement affiché
        if not self.current_worksheet or not self.students:
            return
        
        # Trouver l'étudiant actuellement sélectionné
        current_student = None
        for student_name, worksheets in self.students.items():
            if self.current_worksheet in worksheets:
                current_student = student_name
                break
        
        if not current_student:
            QMessageBox.warning(self, "Erreur", "Aucun étudiant sélectionné")
            return
        
        pdf_path = self.students[current_student][self.current_worksheet]
        
        try:
            # Charger le PDF et la première page
            images = convert_from_path(str(pdf_path), first_page=1, last_page=1, dpi=150)
            if not images:
                QMessageBox.warning(self, "Erreur", "Impossible de charger le PDF")
                return
            
            img = images[0]
            
            # Chercher la position du champ dans field_positions
            field_position = None
            for field_key, position in self.field_positions.items():
                field_info = self.notation_spinboxes.get(field_key)
                if field_info and position and field_name in str(position):
                    field_position = position
                    break
            
            # Si pas de position exacte, chercher par le nom du champ
            if not field_position:
                # Extraire les champs pour trouver la position
                fields = self.extract_pdf_fields(pdf_path)
                for field_key, field_info in fields.items():
                    if field_info.get('name') == field_name:
                        field_position = field_info.get('position')
                        break
            
            # Dessiner un rectangle autour du champ s'il a une position
            if field_position and field_position:
                from PIL import ImageDraw
                
                # Les coordonnées PDF doivent être converties en pixels image
                # pdfplumber/pdf2image utilise une échelle standard
                # Calculer le facteur d'échelle (150 dpi / 72 dpi standard)
                scale = 150 / 72
                
                x0 = int(field_position.get('x0', 0) * scale)
                y0 = int(field_position.get('y0', 0) * scale)
                x1 = int(field_position.get('x1', 0) * scale)
                y1 = int(field_position.get('y1', 0) * scale)
                
                # Créer une copie de l'image pour ne pas modifier l'originale
                img_marked = img.copy()
                draw = ImageDraw.Draw(img_marked, 'RGBA')
                
                # Dessiner un rectangle avec bordure rouge et fond semi-transparent
                draw.rectangle(
                    [(x0-5, y0-5), (x1+5, y1+5)],
                    outline=(255, 0, 0),  # Rouge vif
                    width=3
                )
                
                # Ajouter un surlignage semi-transparent
                draw.rectangle(
                    [(x0, y0), (x1, y1)],
                    fill=(255, 255, 0, 80)  # Jaune semi-transparent
                )
                
                img = img_marked
            
            # Redimensionner pour adapter à la taille du label
            max_width = 350
            max_height = 500
            img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
            
            # Convertir PIL Image en QPixmap via données brutes
            with io.BytesIO() as output:
                img.save(output, format="PNG")
                data = output.getvalue()
            
            pixmap = QPixmap()
            pixmap.loadFromData(data, "PNG")
            
            # Afficher dans la vignette étudiant avec un message
            self.student_pdf_label.setPixmap(pixmap)
            self.student_pdf_label.setAlignment(Qt.AlignCenter)
            
            # Afficher aussi un message
            QMessageBox.information(
                self,
                "Champ localisé",
                f"✅ Champ surligné: <b>{field_name}</b>\n\n"
                f"Le rectangle rouge montre l'emplacement exact du champ.",
                QMessageBox.Ok
            )
        
        except Exception as e:
            print(f"❌ Erreur surbrillance champ: {e}")
            QMessageBox.warning(
                self,
                "Erreur",
                f"Impossible de surbriller le champ: {str(e)}"
            )
    
    def extract_field_values(self, pdf_path):
        """
        Extrait les valeurs saisies dans les champs AcroForm d'un PDF
        Retourne un dictionnaire {field_name: value}
        """
        field_values = {}
        
        try:
            if not PYPDF2_AVAILABLE:
                return field_values
            
            reader = PdfReader(str(pdf_path))
            if not reader.get_fields():
                return field_values
            
            for field_name, field_obj in reader.get_fields().items():
                # Récupérer la valeur du champ
                value = field_obj.get('/V')
                
                # Convertir la valeur en chaîne lisible
                if value:
                    if isinstance(value, bytes):
                        try:
                            value = value.decode('utf-8')
                        except:
                            value = str(value)
                    else:
                        value = str(value)
                else:
                    value = ""
                
                # Simplifier le nom du champ
                simple_name = field_name.split('.')[-1] if '.' in field_name else field_name
                field_values[simple_name] = value
            
            return field_values
        
        except Exception as e:
            print(f"⚠️ Erreur extraction valeurs champs: {e}")
            return field_values
    
    def open_correction_interface(self):
        """Ouvre l'interface de correction comparative"""
        if not self.current_worksheet or not self.students:
            QMessageBox.warning(self, "Erreur", "Veuillez sélectionner un élève d'abord")
            return
        
        # Trouver l'étudiant actuellement sélectionné
        current_student = None
        current_item = self.student_list.currentItem()
        if current_item:
            current_student = current_item.data(Qt.UserRole)
        
        if not current_student or current_student not in self.students:
            QMessageBox.warning(self, "Erreur", "Aucun étudiant sélectionné")
            return
        
        # Vérifier qu'on a les fichiers nécessaires
        student_pdf = self.students[current_student].get(self.current_worksheet)
        if not student_pdf:
            QMessageBox.warning(self, "Erreur", "Fichier étudiant non trouvé")
            return
        
        if not self.correction_dir:
            QMessageBox.warning(self, "Erreur", "Répertoire des corrections non défini")
            return
        
        correction_pdf = Path(self.correction_dir) / f"correction_worksheet{self.current_worksheet}.pdf"
        if not correction_pdf.exists():
            QMessageBox.warning(self, "Erreur", f"Correction non trouvée: {correction_pdf.name}")
            return
        
        # Ouvrir la fenêtre de correction
        correction_dialog = CorrectionDialog(
            student_name=current_student,
            student_pdf=student_pdf,
            correction_pdf=correction_pdf,
            config_manager=self.config_manager,
            parent=self
        )
        correction_dialog.exec_()
    
    def open_batch_correction(self):
        """Ouvre l'interface de correction batch (tous les élèves)"""
        if not self.current_worksheet:
            QMessageBox.warning(
                self, "Erreur", 
                "Veuillez d'abord sélectionner un Worksheet"
            )
            return
        
        if not self.students:
            QMessageBox.warning(self, "Erreur", "Aucun élève chargé")
            return
        
        if not self.correction_dir:
            QMessageBox.warning(self, "Erreur", "Répertoire des corrections non défini")
            return
        
        correction_pdf = Path(self.correction_dir) / f"correction_worksheet{self.current_worksheet}.pdf"
        if not correction_pdf.exists():
            QMessageBox.warning(self, "Erreur", f"Correction non trouvée: {correction_pdf.name}")
            return
        
        # Récupérer la liste des élèves ayant ce worksheet
        eligible_students = [
            name for name, worksheets in self.students.items() 
            if self.current_worksheet in worksheets
        ]
        
        if not eligible_students:
            QMessageBox.warning(
                self, "Erreur", 
                f"Aucun élève n'a de fichier pour Worksheet {self.current_worksheet}"
            )
            return
        
        # Ouvrir la fenêtre de correction batch
        batch_dialog = BatchCorrectionDialog(
            worksheet_num=self.current_worksheet,
            student_names=eligible_students,
            students_dict=self.students,
            correction_pdf=correction_pdf,
            config_manager=self.config_manager,
            parent=self
        )
        batch_dialog.exec_()
    
    def open_cheat_detection(self):
        """Ouvre l'interface de détection de fraude pour un worksheet"""
        if not self.current_worksheet:
            QMessageBox.warning(
                self, "Erreur", 
                "Veuillez d'abord sélectionner un Worksheet"
            )
            return
        
        if not self.students:
            QMessageBox.warning(self, "Erreur", "Aucun élève chargé")
            return
        
        # Vérifier s'il y a des élèves avec ce worksheet
        eligible_students = {
            name: worksheets for name, worksheets in self.students.items()
            if self.current_worksheet in worksheets
        }
        
        if not eligible_students:
            QMessageBox.warning(
                self, "Erreur", 
                f"Aucun élève n'a de fichier pour Worksheet {self.current_worksheet}"
            )
            return
        
        # Ouvrir la fenêtre de détection de fraude
        cheat_detection_dialog = CheatDetectionDialog(
            students_dict=eligible_students,
            worksheet_num=self.current_worksheet,
            parent=self
        )
        cheat_detection_dialog.exec_()


class CorrectionDialog(QDialog):
    """Dialog pour la correction comparative des champs de formulaire"""
    
    def __init__(self, student_name, student_pdf, correction_pdf, config_manager, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Correction - {student_name}")
        self.setGeometry(50, 50, 1400, 700)
        
        self.student_name = student_name
        self.student_pdf = Path(student_pdf)
        self.correction_pdf = Path(correction_pdf)
        self.config_manager = config_manager
        
        # Déterminer le numéro de worksheet avant l'extraction
        self.worksheet_num = None
        for i in range(1, 5):
            if f"worksheet{i}" in str(self.correction_pdf):
                self.worksheet_num = i
                break
        
        # Extraire les valeurs des champs
        self.student_values = self._extract_values(self.student_pdf)
        self.correction_values = self._extract_values(self.correction_pdf)
        
        # Récupérer tous les noms de champs en maintenant l'ordre d'itération du PDF
        all_fields = set(self.student_values.keys()) | set(self.correction_values.keys())
        
        # Utiliser l'ordre du PDF original (plus fiable que de trier par position)
        field_names_ordered = []
        try:
            if PYPDF2_AVAILABLE:
                reader = PdfReader(str(self.student_pdf))
                for field_name in reader.get_fields().keys():
                    simple_name = field_name.split('.')[-1] if '.' in field_name else field_name
                    if simple_name in all_fields and simple_name not in field_names_ordered:
                        field_names_ordered.append(simple_name)
        except:
            pass
        
        # Ajouter les champs manquants
        for f in sorted(all_fields):
            if f not in field_names_ordered:
                field_names_ordered.append(f)
        
        self.field_names = field_names_ordered
        
        self.init_ui()
    
    def _extract_values(self, pdf_path):
        """Extrait les valeurs saisies d'un PDF avec positions"""
        field_data = {}
        try:
            if not PYPDF2_AVAILABLE:
                return field_data
            
            reader = PdfReader(str(pdf_path))
            if not reader.get_fields():
                return field_data
            
            for field_name, field_obj in reader.get_fields().items():
                value = field_obj.get('/V')
                if value:
                    if isinstance(value, bytes):
                        try:
                            value = value.decode('utf-8')
                        except:
                            value = str(value)
                    else:
                        value = str(value)
                else:
                    value = ""
                
                # Extraire position
                position = (999999, 0)
                if '/Rect' in field_obj:
                    try:
                        rect = field_obj['/Rect']
                        y0 = float(rect[1])
                        x0 = float(rect[0])
                        position = (-y0, x0)
                    except:
                        pass
                
                simple_name = field_name.split('.')[-1] if '.' in field_name else field_name
                field_data[simple_name] = {
                    'value': value,
                    'position': position
                }
        
        except Exception as e:
            print(f"⚠️ Erreur extraction valeurs: {e}")
        
        return field_data
    
    def _sort_fields_by_position(self, pdf_path, field_names):
        """
        Trie les champs par leur position dans le PDF
        (de haut en bas, puis de gauche à droite)
        """
        field_positions = {}
        
        try:
            if not PYPDF2_AVAILABLE:
                return sorted(field_names)
            
            reader = PdfReader(str(pdf_path))
            pdf_fields = reader.get_fields()
            
            if not pdf_fields:
                return sorted(field_names)
            
            # Créer un mapping simple_name -> position
            for field_name, field_obj in pdf_fields.items():
                simple_name = field_name.split('.')[-1] if '.' in field_name else field_name
                
                position = (999999, 0)
                try:
                    if '/Rect' in field_obj:
                        rect = field_obj['/Rect']
                        y0 = float(rect[1])
                        x0 = float(rect[0])
                        position = (-y0, x0)
                except:
                    pass
                
                field_positions[simple_name] = position
            
            # Trier
            sorted_fields = sorted(
                field_names,
                key=lambda f: field_positions.get(f, (999999, 0))
            )
            
            return sorted_fields
        
        except Exception as e:
            print(f"⚠️ Erreur tri par position: {e}")
            return sorted(field_names)
    
    def init_ui(self):
        """Initialise l'interface de correction"""
        layout = QVBoxLayout()
        
        # Titre avec nom de l'élève
        title_label = QLabel(f"Correction - {self.student_name}")
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title_label.setFont(title_font)
        layout.addWidget(title_label)
        
        # Tableau de comparaison
        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels([
            "N°",
            "Réponse étudiant",
            "Réponse correction",
            "Note",
            "Points/champ"
        ])
        
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        
        table.setRowCount(len(self.field_names))
        
        # Dictionnaire pour stocker les spinbox
        self.field_spinboxes = {}
        self.field_names_map = {}  # Stocker les données pour affichage complet au clic
        
        # Label pour afficher le score total
        self.total_score_label = QLabel()
        self.total_score_label.setStyleSheet(
            "font-size: 14px; font-weight: bold; padding: 10px; background-color: #F0F0F0; border-radius: 5px;"
        )
        
        # Importer la fonction de normalisation
        try:
            from correction import normalize_text
        except ImportError:
            def normalize_text(text):
                """Fallback: normalisation simple si correction.py n'est pas disponible"""
                return text.lower().strip() if text else ""
        
        for row, field_name in enumerate(self.field_names):
            student_data = self.student_values.get(field_name, {})
            correction_data = self.correction_values.get(field_name, {})
            
            # Extraire la valeur (compatible dict ou string)
            if isinstance(student_data, dict):
                student_value = student_data.get('value', '')
            else:
                student_value = student_data or ''
            
            if isinstance(correction_data, dict):
                correction_value = correction_data.get('value', '')
            else:
                correction_value = correction_data or ''
            
            # Normaliser les valeurs pour la comparaison
            student_normalized = normalize_text(student_value)
            correction_normalized = normalize_text(correction_value)
            
            # Numéro du champ
            num_item = QTableWidgetItem(str(row + 1))
            num_item.setFlags(num_item.flags() & ~Qt.ItemIsEditable)
            table.setItem(row, 0, num_item)
            
            # Réponse étudiant - afficher la valeur normalisée
            student_item = QTableWidgetItem(student_normalized if student_normalized else "[VIDE]")
            student_item.setFlags(student_item.flags() & ~Qt.ItemIsEditable)
            
            # Colorer selon si la réponse est correcte (comparaison normalisée)
            is_correct = student_normalized and student_normalized == correction_normalized
            if not student_normalized:
                student_item.setBackground(QColor(255, 200, 200))  # Rouge
            elif is_correct:
                student_item.setBackground(QColor(200, 255, 200))  # Vert
            else:
                student_item.setBackground(QColor(255, 200, 200))  # Rouge
            
            table.setItem(row, 1, student_item)
            
            # Réponse correction - afficher la valeur normalisée
            correction_item = QTableWidgetItem(correction_normalized)
            correction_item.setFlags(correction_item.flags() & ~Qt.ItemIsEditable)
            correction_item.setBackground(QColor(200, 255, 200))
            table.setItem(row, 2, correction_item)
            
            # === COLONNE 3: Note obtenue par l'étudiant (EDITABLE) ===
            # Charger le barème
            barème_points = self._load_barème_point(field_name)
            if barème_points is None:
                barème_points = 1
            
            note_value = barème_points if is_correct else 0
            
            # Créer une spinbox éditable pour la note
            note_spinbox = QSpinBox()
            note_spinbox.setMinimum(0)
            note_spinbox.setMaximum(barème_points)  # Le maximum de note = le barème
            note_spinbox.setValue(note_value)
            note_spinbox.valueChanged.connect(self._update_total_score)
            
            table.setCellWidget(row, 3, note_spinbox)
            
            # === COLONNE 4: Points par champ (barème - EDITABLE) ===
            barème_spinbox = QSpinBox()
            barème_spinbox.setMinimum(0)
            barème_spinbox.setMaximum(999)
            barème_spinbox.setValue(barème_points)
            
            barème_spinbox.valueChanged.connect(
                lambda val, fn=field_name: self._on_barème_changed(fn, val)
            )
            barème_spinbox.valueChanged.connect(self._update_total_score)
            
            table.setCellWidget(row, 4, barème_spinbox)
            
            # Stocker les spinboxes
            self.field_spinboxes[field_name] = {
                'note_spinbox': note_spinbox,
                'barème_spinbox': barème_spinbox,
                'is_correct': is_correct
            }
            
            # Stocker les données pour affichage complet au clic (avec valeurs normalisées)
            self.field_names_map = getattr(self, 'field_names_map', {})
            self.field_names_map[row] = {
                'field_name': field_name,
                'student_value': student_normalized,
                'correction_value': correction_normalized
            }
        
        layout.addWidget(table)
        
        # Connecter le clic sur le tableau
        table.cellClicked.connect(self._on_table_cell_clicked)
        
        # Ajouter le label de score total
        layout.addWidget(self.total_score_label)
        
        # Mettre à jour le score total
        self._update_total_score()
        
        # Boutons d'action
        button_layout = QHBoxLayout()
        
        save_button = QPushButton("✓ Enregistrer les points")
        save_button.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")
        save_button.clicked.connect(self.save_scores)
        button_layout.addWidget(save_button)
        
        cancel_button = QPushButton("Annuler")
        cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(cancel_button)
        
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def save_scores(self):
        """Sauvegarde les points attribués"""
        if not self.worksheet_num:
            QMessageBox.warning(self, "Erreur", "Impossible de déterminer le worksheet")
            return
        
        # Sauvegarder les points
        if "student_scores" not in self.config_manager.config:
            self.config_manager.config["student_scores"] = {}
        
        if self.student_name not in self.config_manager.config["student_scores"]:
            self.config_manager.config["student_scores"][self.student_name] = {}
        
        ws_key = f"worksheet{self.worksheet_num}"
        if ws_key not in self.config_manager.config["student_scores"][self.student_name]:
            self.config_manager.config["student_scores"][self.student_name][ws_key] = {}
        
        # Sauvegarder les scores pour chaque champ
        total_score = 0
        max_score = 0
        for field_name, data in self.field_spinboxes.items():
            note_value = data['note_spinbox'].value()
            barème_value = data['barème_spinbox'].value()
            
            field_key = f"field_{field_name}"
            self.config_manager.config["student_scores"][self.student_name][ws_key][field_key] = note_value
            
            total_score += note_value
            max_score += barème_value
        
        self.config_manager.save_config()
        
        QMessageBox.information(
            self,
            "Succès",
            f"✓ Points enregistrés\n\n"
            f"Note totale: {total_score} / {max_score}"
        )
        
        self.accept()
    
    def _on_table_cell_clicked(self, row, col):
        """Affiche le contenu complet d'une cellule au clic (colonnes 1 et 2)"""
        if col not in [1, 2]:  # Colonnes "Réponse étudiant" et "Réponse correction"
            return
        
        # Récupérer les données du champ
        field_data = self.field_names_map.get(row)
        if not field_data:
            return
        
        # Ouvrir le dialogue
        dialog = FieldTextDialog(
            field_data['field_name'],
            field_data['student_value'],
            field_data['correction_value'],
            parent=self
        )
        dialog.exec_()
    
    def _load_barème_point(self, field_name):
        """Charge le barème de points pour un champ"""
        barème_key = f"field_bareme_{field_name}"
        if "bareme_points" not in self.config_manager.config:
            self.config_manager.config["bareme_points"] = {}
        
        ws_key = f"worksheet{self.worksheet_num}"
        if ws_key not in self.config_manager.config["bareme_points"]:
            self.config_manager.config["bareme_points"][ws_key] = {}
        
        return self.config_manager.config["bareme_points"][ws_key].get(barème_key)
    
    def _on_barème_changed(self, field_name, value):
        """Gère le changement de barème et sauvegarde"""
        barème_key = f"field_bareme_{field_name}"
        if "bareme_points" not in self.config_manager.config:
            self.config_manager.config["bareme_points"] = {}
        
        ws_key = f"worksheet{self.worksheet_num}"
        if ws_key not in self.config_manager.config["bareme_points"]:
            self.config_manager.config["bareme_points"][ws_key] = {}
        
        self.config_manager.config["bareme_points"][ws_key][barème_key] = value
        self.config_manager.save_config()
        
        # Mettre à jour le maximum de la note pour qu'elle ne dépasse pas le barème
        note_spinbox = self.field_spinboxes[field_name]['note_spinbox']
        note_spinbox.setMaximum(value)
        # Si la note actuelle dépasse le nouveau barème, la réduire
        if note_spinbox.value() > value:
            note_spinbox.setValue(value)
    
    def _update_total_score(self):
        """Met à jour l'affichage du score total"""
        total_score = 0
        max_score = 0
        
        for field_name, data in self.field_spinboxes.items():
            note_value = data['note_spinbox'].value()
            barème_value = data['barème_spinbox'].value()
            
            total_score += note_value
            max_score += barème_value
        
        # Calculer le pourcentage
        percentage = (total_score / max_score * 100) if max_score > 0 else 0
        
        self.total_score_label.setText(
            f"📊 Note totale : <b>{total_score} / {max_score}</b> ({percentage:.1f}%)"
        )


class BatchCorrectionDialog(QDialog):
    """Dialog pour la correction comparative en batch de plusieurs élèves"""
    
    def __init__(self, worksheet_num, student_names, students_dict, correction_pdf, config_manager, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Correction Batch - Worksheet {worksheet_num}")
        self.setGeometry(50, 50, 1500, 800)
        
        self.worksheet_num = worksheet_num
        self.student_names = sorted(student_names)
        self.students_dict = students_dict
        self.correction_pdf = Path(correction_pdf)
        self.config_manager = config_manager
        
        self.current_student_index = 0
        self.correction_data = {}  # Stockage des corrections par élève {student: {field: score}}
        
        # Extraire les valeurs de correction une seule fois
        self.correction_values = self._extract_values(self.correction_pdf)
        
        self.init_ui()
        self.load_current_student()
    
    def _extract_values(self, pdf_path):
        """Extrait les valeurs saisies d'un PDF avec positions"""
        field_data = {}
        try:
            if not PYPDF2_AVAILABLE:
                return field_data
            
            reader = PdfReader(str(pdf_path))
            if not reader.get_fields():
                return field_data
            
            for field_name, field_obj in reader.get_fields().items():
                value = field_obj.get('/V')
                if value:
                    if isinstance(value, bytes):
                        try:
                            value = value.decode('utf-8')
                        except:
                            value = str(value)
                    else:
                        value = str(value)
                else:
                    value = ""
                
                # Extraire position
                position = (999999, 0)
                if '/Rect' in field_obj:
                    try:
                        rect = field_obj['/Rect']
                        y0 = float(rect[1])
                        x0 = float(rect[0])
                        position = (-y0, x0)
                    except:
                        pass
                
                simple_name = field_name.split('.')[-1] if '.' in field_name else field_name
                field_data[simple_name] = {
                    'value': value,
                    'position': position
                }
        
        except Exception as e:
            print(f"⚠️ Erreur extraction valeurs: {e}")
        
        return field_data
    
    def _sort_fields_by_position(self, pdf_path, field_names):
        """
        Trie les champs par leur position dans le PDF
        (de haut en bas, puis de gauche à droite)
        Extrait directement les positions du PDF original.
        """
        field_positions = {}
        
        try:
            if not PYPDF2_AVAILABLE:
                return sorted(field_names)
            
            reader = PdfReader(str(pdf_path))
            pdf_fields = reader.get_fields()
            
            if not pdf_fields:
                return sorted(field_names)
            
            # Créer un mapping simple_name -> field_name complet et position
            name_to_info = {}
            for field_name, field_obj in pdf_fields.items():
                simple_name = field_name.split('.')[-1] if '.' in field_name else field_name
                
                # Récupérer la position du champ
                position = (999999, 0)  # Par défaut: très bas, très à gauche (fin de liste)
                
                try:
                    if '/Rect' in field_obj:
                        rect = field_obj['/Rect']
                        # rect = [x0, y0, x1, y1]
                        # Y en PDF croît vers le haut (0 en bas, 842 en haut pour A4)
                        # On veut trier de haut en bas, donc on utilise -y0
                        y0 = float(rect[1])
                        x0 = float(rect[0])
                        # Trier par: -y0 (haut en bas), puis x0 (gauche à droite)
                        position = (-y0, x0)
                except Exception as e:
                    print(f"⚠️ Erreur extraction rect pour {field_name}: {e}")
                
                name_to_info[simple_name] = position
            
            # Trier les champs demandés par position
            sorted_fields = sorted(
                field_names,
                key=lambda f: name_to_info.get(f, (999999, 0))
            )
            
            print(f"📍 Ordre des champs par position:")
            for i, f in enumerate(sorted_fields):
                pos = name_to_info.get(f, (999999, 0))
                print(f"  {i+1}. {f} → y={-pos[0]:.0f}, x={pos[1]:.0f}")
            
            return sorted_fields
        
        except Exception as e:
            print(f"⚠️ Erreur tri par position: {e}")
            import traceback
            traceback.print_exc()
            return sorted(field_names)
    
    def _update_score_display(self):
        """Met à jour l'affichage de la note totale basée sur les notes entrées"""
        if not hasattr(self, 'field_spinboxes'):
            return
        
        total = 0
        max_possible = 0
        
        for field_name, data in self.field_spinboxes.items():
            note_value = data['note_spinbox'].value()
            barème_points = data['barème_spinbox'].value()
            
            total += note_value
            max_possible += barème_points
        
        if max_possible > 0:
            self.score_label.setText(f"Note: {total} / {max_possible}")
        else:
            self.score_label.setText("Note: 0 / 0")
    
    def init_ui(self):
        """Initialise l'interface de correction batch"""
        main_layout = QHBoxLayout()
        
        # === COLONNE GAUCHE : LISTE DES ÉLÈVES ===
        left_layout = QVBoxLayout()
        left_layout.addWidget(QLabel("Élèves à corriger:"))
        
        self.student_list_widget = QListWidget()
        self.student_list_widget.itemClicked.connect(self.on_student_list_clicked)
        
        for i, name in enumerate(self.student_names):
            item = QListWidgetItem(name)
            item.setData(Qt.UserRole, i)
            self.student_list_widget.addItem(item)
        
        left_layout.addWidget(self.student_list_widget)
        
        # Statistiques
        stats_label = QLabel()
        stats_label.setText(f"Total: {len(self.student_names)} élèves")
        stats_label.setStyleSheet("font-weight: bold; color: blue;")
        left_layout.addWidget(stats_label)
        self.stats_label = stats_label
        
        left_widget = QWidget()
        left_widget.setLayout(left_layout)
        left_widget.setMaximumWidth(200)
        main_layout.addWidget(left_widget)
        
        # === COLONNE DROITE : INTERFACE DE CORRECTION ===
        right_layout = QVBoxLayout()
        
        # Titre avec info élève
        title_layout = QHBoxLayout()
        self.title_label = QLabel()
        self.title_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        title_layout.addWidget(self.title_label)
        title_layout.addStretch()
        right_layout.addLayout(title_layout)
        
        # Tableau de comparaison
        self.correction_table = QTableWidget()
        self.correction_table.setColumnCount(5)
        self.correction_table.setHorizontalHeaderLabels([
            "N°",
            "Réponse étudiant",
            "Réponse correction",
            "Note",
            "Points/champ"
        ])
        
        self.correction_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.correction_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.correction_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.correction_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.correction_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        
        # Connecter le clic sur le tableau pour voir le texte complet
        self.correction_table.cellClicked.connect(self._on_batch_table_cell_clicked)
        
        right_layout.addWidget(self.correction_table)
        
        # Label pour afficher la note totale
        self.score_label = QLabel("Note: 0 / 0")
        self.score_label.setStyleSheet("font-weight: bold; font-size: 12px; color: blue; padding: 5px;")
        self.score_label.setAlignment(Qt.AlignRight)
        right_layout.addWidget(self.score_label)
        
        # Boutons de navigation
        nav_layout = QHBoxLayout()
        
        self.btn_prev = QPushButton("◀ Précédent")
        self.btn_prev.clicked.connect(self.prev_student)
        nav_layout.addWidget(self.btn_prev)
        
        self.progress_label = QLabel()
        self.progress_label.setAlignment(Qt.AlignCenter)
        self.progress_label.setStyleSheet("font-weight: bold;")
        nav_layout.addWidget(self.progress_label)
        
        self.btn_next = QPushButton("Suivant ▶")
        self.btn_next.clicked.connect(self.next_student)
        nav_layout.addWidget(self.btn_next)
        
        right_layout.addLayout(nav_layout)
        
        # Boutons d'action
        action_layout = QHBoxLayout()
        
        btn_save_all = QPushButton("✓ Enregistrer TOUT et fermer")
        btn_save_all.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")
        btn_save_all.clicked.connect(self.save_all_and_close)
        action_layout.addWidget(btn_save_all)
        
        btn_export = QPushButton("📊 Exporter vers Calc")
        btn_export.setStyleSheet("background-color: #2196F3; color: white; font-weight: bold;")
        btn_export.clicked.connect(self.export_results_to_spreadsheet)
        action_layout.addWidget(btn_export)
        
        btn_cancel = QPushButton("Annuler")
        btn_cancel.clicked.connect(self.reject)
        action_layout.addWidget(btn_cancel)
        
        right_layout.addLayout(action_layout)
        
        right_widget = QWidget()
        right_widget.setLayout(right_layout)
        main_layout.addWidget(right_widget)
        
        self.setLayout(main_layout)
    
    def load_current_student(self):
        """Charge et affiche le formulaire de l'élève courant"""
        if self.current_student_index >= len(self.student_names):
            return
        
        # Importer la fonction de normalisation
        try:
            from correction import normalize_text
        except ImportError:
            def normalize_text(text):
                """Fallback: normalisation simple si correction.py n'est pas disponible"""
                return text.lower().strip() if text else ""
        
        # Sauvegarder les données de l'élève précédent
        if self.current_student_index > 0:
            prev_student = self.student_names[self.current_student_index - 1]
            self._save_current_student_data(prev_student)
        
        current_student = self.student_names[self.current_student_index]
        
        # Mise à jour du titre
        self.title_label.setText(
            f"Élève {self.current_student_index + 1}/{len(self.student_names)} - {current_student}"
        )
        
        # Mise à jour de la barre de progression
        self.progress_label.setText(f"{self.current_student_index + 1} / {len(self.student_names)}")
        
        # Activer/désactiver les boutons de navigation
        self.btn_prev.setEnabled(self.current_student_index > 0)
        self.btn_next.setEnabled(self.current_student_index < len(self.student_names) - 1)
        
        # Sélectionner dans la liste
        self.student_list_widget.setCurrentRow(self.current_student_index)
        
        # Extraire les valeurs du PDF étudiant
        student_pdf = self.students_dict[current_student].get(self.worksheet_num)
        student_values = self._extract_values(student_pdf)
        
        # Récupérer tous les noms de champs en maintenant l'ordre d'itération du PDF
        all_fields = set(student_values.keys()) | set(self.correction_values.keys())
        
        # Utiliser l'ordre du PDF original (plus fiable que de trier par position)
        field_names_ordered = []
        try:
            if PYPDF2_AVAILABLE:
                reader = PdfReader(str(student_pdf))
                for field_name in reader.get_fields().keys():
                    simple_name = field_name.split('.')[-1] if '.' in field_name else field_name
                    if simple_name in all_fields and simple_name not in field_names_ordered:
                        field_names_ordered.append(simple_name)
        except:
            pass
        
        # Ajouter les champs manquants
        for f in sorted(all_fields):
            if f not in field_names_ordered:
                field_names_ordered.append(f)
        
        field_names = field_names_ordered
        
        # Remplir le tableau
        self.correction_table.setRowCount(len(field_names))
        self.field_spinboxes = {}
        self.batch_field_data = {}  # Stocker les données pour affichage complet au clic
        
        for row, field_name in enumerate(field_names):
            student_data = student_values.get(field_name, {})
            correction_data = self.correction_values.get(field_name, {})
            
            # Extraire la valeur (compatible dict ou string)
            if isinstance(student_data, dict):
                student_value = student_data.get('value', '')
            else:
                student_value = student_data or ''
            
            if isinstance(correction_data, dict):
                correction_value = correction_data.get('value', '')
            else:
                correction_value = correction_data or ''
            
            # Normaliser les valeurs pour la comparaison
            student_normalized = normalize_text(student_value)
            correction_normalized = normalize_text(correction_value)
            
            # Numéro
            num_item = QTableWidgetItem(str(row + 1))
            num_item.setFlags(num_item.flags() & ~Qt.ItemIsEditable)
            self.correction_table.setItem(row, 0, num_item)
            
            # Réponse étudiant - afficher la valeur normalisée
            student_item = QTableWidgetItem(student_normalized if student_normalized else "[VIDE]")
            student_item.setFlags(student_item.flags() & ~Qt.ItemIsEditable)
            
            # Colorer selon si la réponse est correcte (comparaison normalisée)
            if not student_normalized:
                student_item.setBackground(QColor(255, 200, 200))  # Rouge
            elif student_normalized == correction_normalized:
                student_item.setBackground(QColor(200, 255, 200))  # Vert
            else:
                student_item.setBackground(QColor(255, 200, 200))  # Rouge
            
            self.correction_table.setItem(row, 1, student_item)
            
            # Réponse correction - afficher la valeur normalisée
            correction_item = QTableWidgetItem(correction_normalized)
            correction_item.setFlags(correction_item.flags() & ~Qt.ItemIsEditable)
            correction_item.setBackground(QColor(200, 255, 200))
            self.correction_table.setItem(row, 2, correction_item)
            
            # === COLONNE 3: Note obtenue par l'étudiant (EDITABLE) ===
            is_correct = student_normalized and student_normalized == correction_normalized
            barème_points = self._load_barème_point(field_name)
            if barème_points is None:
                barème_points = 1 if is_correct else 1
            
            note_value = barème_points if is_correct else 0
            
            # Créer une spinbox éditable pour la note
            note_spinbox = QSpinBox()
            note_spinbox.setMinimum(0)
            note_spinbox.setMaximum(barème_points)  # Le maximum de note = le barème
            note_spinbox.setValue(note_value)
            
            note_spinbox.valueChanged.connect(
                lambda val, fn=field_name: self._on_note_changed(fn, val)
            )
            
            self.correction_table.setCellWidget(row, 3, note_spinbox)
            
            # === COLONNE 4: Points par champ (barème - EDITABLE) ===
            barème_spinbox = QSpinBox()
            barème_spinbox.setMinimum(0)
            barème_spinbox.setMaximum(999)
            barème_spinbox.setValue(barème_points)
            
            barème_spinbox.valueChanged.connect(
                lambda val, fn=field_name: self._on_barème_changed(fn, val)
            )
            
            self.correction_table.setCellWidget(row, 4, barème_spinbox)
            
            # Stocker les spinboxes et le field_name pour mise à jour
            self.field_spinboxes[field_name] = {
                'note_spinbox': note_spinbox,
                'barème_spinbox': barème_spinbox,
                'is_correct': is_correct
            }
            
            # Stocker les données pour affichage complet au clic (avec valeurs normalisées)
            self.batch_field_data[row] = {
                'field_name': field_name,
                'student_value': student_normalized,
                'correction_value': correction_normalized
            }
        
        # Afficher la note
        self._update_score_display()
    
    def _load_saved_score(self, student_name, field_key):
        """Charge un score sauvegardé depuis la config"""
        if "student_scores" not in self.config_manager.config:
            return None
        
        if student_name not in self.config_manager.config["student_scores"]:
            return None
        
        ws_key = f"worksheet{self.worksheet_num}"
        if ws_key not in self.config_manager.config["student_scores"][student_name]:
            return None
        
        scores = self.config_manager.config["student_scores"][student_name][ws_key]
        return scores.get(field_key)
    
    def _save_current_student_data(self, student_name):
        """Sauvegarde les données de correction pour un élève"""
        if not hasattr(self, 'field_spinboxes'):
            return
        
        if student_name not in self.correction_data:
            self.correction_data[student_name] = {}
        
        for field_name, data in self.field_spinboxes.items():
            field_key = f"field_{field_name}"
            # Sauvegarder la note obtenue (valeur de la spinbox Note)
            note_value = data['note_spinbox'].value()
            self.correction_data[student_name][field_key] = note_value
    
    def _load_barème_point(self, field_name):
        """Charge le barème de points pour un champ"""
        barème_key = f"field_bareme_{field_name}"
        if "bareme_points" not in self.config_manager.config:
            self.config_manager.config["bareme_points"] = {}
        
        ws_key = f"worksheet{self.worksheet_num}"
        if ws_key not in self.config_manager.config["bareme_points"]:
            self.config_manager.config["bareme_points"][ws_key] = {}
        
        return self.config_manager.config["bareme_points"][ws_key].get(barème_key)
    
    def _on_note_changed(self, field_name, value):
        """Gère le changement de note et met à jour l'affichage du score total"""
        # Mettre à jour le score total en temps réel
        self._update_score_display()
    
    def _on_barème_changed(self, field_name, value):
        """Gère le changement de barème, sauvegarde et met à jour le score"""
        # Sauvegarder le barème en config
        barème_key = f"field_bareme_{field_name}"
        if "bareme_points" not in self.config_manager.config:
            self.config_manager.config["bareme_points"] = {}
        
        ws_key = f"worksheet{self.worksheet_num}"
        if ws_key not in self.config_manager.config["bareme_points"]:
            self.config_manager.config["bareme_points"][ws_key] = {}
        
        self.config_manager.config["bareme_points"][ws_key][barème_key] = value
        self.config_manager.save_config()
        
        # Mettre à jour le maximum de la note pour qu'elle ne dépasse pas le barème
        note_spinbox = self.field_spinboxes[field_name]['note_spinbox']
        note_spinbox.setMaximum(value)
        # Si la note actuelle dépasse le nouveau barème, la réduire
        if note_spinbox.value() > value:
            note_spinbox.setValue(value)
        
        # Mettre à jour le score total en temps réel
        self._update_score_display()
    
    def _on_field_note_changed(self, field_name, note_value):
        """Callback appelé quand une note change (actuellement unused, peut être utilisé pour des mises à jour spéciales)"""
        pass
    
    def on_student_list_clicked(self, item):
        """Gère le clic sur un élève dans la liste"""
        index = item.data(Qt.UserRole)
        if index != self.current_student_index:
            self.current_student_index = index
            self.load_current_student()
    
    def prev_student(self):
        """Va à l'élève précédent"""
        if self.current_student_index > 0:
            self.current_student_index -= 1
            self.load_current_student()
    
    def next_student(self):
        """Va à l'élève suivant"""
        if self.current_student_index < len(self.student_names) - 1:
            self.current_student_index += 1
            self.load_current_student()
    
    def export_results_to_spreadsheet(self):
        """Exporte les résultats de correction vers un onglet 'récap' du fichier Calc"""
        # Sauvegarder d'abord l'élève courant
        current_student = self.student_names[self.current_student_index]
        self._save_current_student_data(current_student)
        
        # Récupérer le chemin du dernier fichier Calc utilisé
        spreadsheet_path = self.config_manager.config.get("last_spreadsheet_path")
        
        if not spreadsheet_path or not Path(spreadsheet_path).exists():
            QMessageBox.warning(self, "Aucun fichier de planning",
                "Veuillez d'abord importer un fichier de planning (Calc/Excel) depuis la fenêtre principale.")
            return
        
        try:
            from openpyxl import load_workbook
            from odf.opendocument import load as odf_load
            from odf.table import Table
            
            # Déterminer le format du fichier
            file_ext = Path(spreadsheet_path).suffix.lower()
            
            if file_ext == '.xlsx':
                # Traiter le fichier Excel
                workbook = load_workbook(spreadsheet_path)
                
                # Créer ou récupérer l'onglet "récap"
                if "récap" in workbook.sheetnames:
                    ws = workbook["récap"]
                    ws.delete_rows(2, ws.max_row)  # Garder l'en-tête, supprimer les données
                else:
                    ws = workbook.create_sheet("récap", 0)
                    # Créer l'en-tête
                    ws['A1'] = 'Étudiant'
                    ws['B1'] = f'Worksheet {self.worksheet_num}'
                
                # Ajouter les résultats
                row = 2
                for student_name, scores_dict in self.correction_data.items():
                    ws[f'A{row}'] = student_name
                    total_score = sum(scores_dict.values())
                    ws[f'B{row}'] = total_score
                    row += 1
                
                workbook.save(spreadsheet_path)
                QMessageBox.information(self, "✓ Export réussi",
                    f"Résultats exportés vers l'onglet 'récap':\n{Path(spreadsheet_path).name}")
                print(f"✅ Résultats exportés vers {spreadsheet_path}")
                
            elif file_ext == '.ods':
                # Traiter le fichier ODS avec odfpy
                from odf.table import Table, TableRow, TableCell
                from odf.text import P
                
                odf_doc = odf_load(spreadsheet_path)
                
                # Trouver ou créer l'onglet "récap"
                tables = odf_doc.spreadsheet.getElementsByType(Table)
                recap_table = None
                
                for table in tables:
                    if table.getAttribute('name') == 'récap':
                        recap_table = table
                        # Supprimer les lignes existantes (garder juste l'en-tête)
                        rows = recap_table.getElementsByType(TableRow)
                        for i in range(len(rows) - 1, 0, -1):  # Garder la première (en-tête)
                            recap_table.removeChild(rows[i])
                        break
                
                if recap_table is None:
                    # Créer une nouvelle table
                    recap_table = Table(name="récap")
                    odf_doc.spreadsheet.addElement(recap_table)
                    
                    # Ajouter l'en-tête
                    header_row = TableRow()
                    recap_table.addElement(header_row)
                    
                    # Cellule 1: "Étudiant"
                    cell1 = TableCell(valuetype="string")
                    p1 = P(text="Étudiant")
                    cell1.addElement(p1)
                    header_row.addElement(cell1)
                    
                    # Cellule 2: "Worksheet X"
                    cell2 = TableCell(valuetype="string")
                    p2 = P(text=f"Worksheet {self.worksheet_num}")
                    cell2.addElement(p2)
                    header_row.addElement(cell2)
                else:
                    # S'assurer que les en-têtes existent
                    rows = recap_table.getElementsByType(TableRow)
                    if len(rows) == 0:
                        # Créer l'en-tête s'il n'existe pas
                        header_row = TableRow()
                        recap_table.addElement(header_row)
                        
                        cell1 = TableCell(valuetype="string")
                        p1 = P(text="Étudiant")
                        cell1.addElement(p1)
                        header_row.addElement(cell1)
                        
                        cell2 = TableCell(valuetype="string")
                        p2 = P(text=f"Worksheet {self.worksheet_num}")
                        cell2.addElement(p2)
                        header_row.addElement(cell2)
                
                # Ajouter les résultats
                for student_name, scores_dict in self.correction_data.items():
                    row = TableRow()
                    recap_table.addElement(row)
                    
                    # Cellule: nom étudiant
                    cell_name = TableCell(valuetype="string")
                    p_name = P(text=student_name)
                    cell_name.addElement(p_name)
                    row.addElement(cell_name)
                    
                    # Cellule: score total
                    total_score = sum(scores_dict.values())
                    cell_score = TableCell(valuetype="float", value=str(total_score))
                    p_score = P(text=str(total_score))
                    cell_score.addElement(p_score)
                    row.addElement(cell_score)
                
                # Sauvegarder le document
                odf_doc.save(spreadsheet_path)
                QMessageBox.information(self, "✓ Export réussi",
                    f"Résultats exportés vers l'onglet 'récap':\n{Path(spreadsheet_path).name}")
                print(f"✅ Résultats exportés vers {spreadsheet_path}")
            else:
                QMessageBox.warning(self, "Format non supporté",
                    "Seuls les fichiers .xlsx (Excel) et .ods (Calc) sont supportés pour l'export automatique.")
        
        except ImportError:
            QMessageBox.critical(self, "Erreur de dépendance",
                "Les bibliothèques openpyxl/odfpy ne sont pas installées.\nInstallation: pip install openpyxl odfpy")
        except Exception as e:
            QMessageBox.critical(self, "Erreur lors de l'export",
                f"Une erreur s'est produite:\n{str(e)}")
            print(f"❌ Erreur export: {e}")
            import traceback
            traceback.print_exc()
    
    def save_all_and_close(self):
        """Sauvegarde tous les scores et ferme la fenêtre"""
        # Sauvegarder l'élève courant
        current_student = self.student_names[self.current_student_index]
        self._save_current_student_data(current_student)
        
        # Sauvegarder tous les scores dans la config
        if "student_scores" not in self.config_manager.config:
            self.config_manager.config["student_scores"] = {}
        
        for student_name, scores_dict in self.correction_data.items():
            if student_name not in self.config_manager.config["student_scores"]:
                self.config_manager.config["student_scores"][student_name] = {}
            
            ws_key = f"worksheet{self.worksheet_num}"
            if ws_key not in self.config_manager.config["student_scores"][student_name]:
                self.config_manager.config["student_scores"][student_name][ws_key] = {}
            
            self.config_manager.config["student_scores"][student_name][ws_key].update(scores_dict)
        
        self.config_manager.save_config()
        
        # Afficher un résumé
        total_students = len(self.student_names)
        avg_score = 0
        total_points = 0
        total_max = 0
        
        for student_name, scores_dict in self.correction_data.items():
            total_points += sum(scores_dict.values())
            total_max += len(scores_dict)
        
        if total_max > 0:
            avg_score = total_points / total_max
        
        QMessageBox.information(
            self,
            "✓ Correction batch terminée",
            f"Tous les scores ont été enregistrés!\n\n"
            f"Élèves corrigés: {total_students}\n"
            f"Points totaux: {total_points} / {total_max}\n"
            f"Moyenne par champ: {avg_score:.2f}"
        )
        
        self.accept()
    
    def _on_batch_table_cell_clicked(self, row, column):
        """Affiche le contenu complet d'une cellule quand on clique dessus"""
        item = self.correction_table.item(row, column)
        if item:
            cell_text = item.text()
            if cell_text and len(cell_text) > 100:  # Seulement si le texte est long
                # Affiche le texte complet dans une boîte de dialogue
                QMessageBox.information(
                    self,
                    f"Contenu complet (ligne {row + 1})",
                    cell_text
                )


def main():
    app = QApplication(sys.argv)
    window = CorrectionApp()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
